"""Compute REF gross/NAV ratio and recommend target leverage."""
from openpyxl import load_workbook
import statistics as st

REF_P = r"C:\Users\werdn\Downloads\DC ETF Arb Attribution.xlsx"
NEW_P = r"C:\Users\werdn\Downloads\DC ETF Arb Pairwise Backtest Attribution.xlsx"

def gather(p):
    wb = load_workbook(p, data_only=True)
    ws = wb["Monthly Attribution"]
    rows = []
    cum = 10_000_000.0
    for r in range(4, ws.max_row + 1):
        m = ws[f"B{r}"].value
        f_long = ws[f"F{r}"].value
        g_short = ws[f"G{r}"].value
        net = ws[f"T{r}"].value
        if m is None or not isinstance(f_long, (int, float)):
            continue
        if isinstance(net, (int, float)):
            cum += net
        s_abs = abs(g_short) if isinstance(g_short, (int, float)) else 0.0
        rows.append({
            "month": m,
            "nav": cum,
            "long": f_long,
            "short_abs": s_abs,
            "gross": f_long + s_abs,
            "long_to_nav": f_long / cum,
            "short_to_nav": s_abs / cum,
            "gross_to_nav": (f_long + s_abs) / cum,
        })
    return rows

ref = gather(REF_P)
new = gather(NEW_P)

def stats(rows, key):
    vals = [r[key] for r in rows]
    return min(vals), max(vals), st.mean(vals), st.median(vals)

print(f"{'metric':>14}  {'REF min':>8} {'REF max':>8} {'REF mean':>9} {'REF med':>9}  |  {'NEW min':>8} {'NEW max':>8} {'NEW mean':>9} {'NEW med':>9}")
for k in ("long_to_nav", "short_to_nav", "gross_to_nav"):
    rmin, rmax, rmean, rmed = stats(ref, k)
    nmin, nmax, nmean, nmed = stats(new, k)
    print(f"{k:>14}  {rmin:>8.2f} {rmax:>8.2f} {rmean:>9.3f} {rmed:>9.3f}  |  {nmin:>8.2f} {nmax:>8.2f} {nmean:>9.3f} {nmed:>9.3f}")

print()
# Implied beta-weighted hedge ratio: short/long
print(f"{'period':>10}  REF L/NAV  REF S/NAV  REF S/L   NEW L/NAV  NEW S/NAV  NEW S/L")
for r, n in zip(ref, new):
    if r["month"] != n["month"]:
        continue
    sl_ref = r["short_abs"] / r["long"] if r["long"] else 0
    sl_new = n["short_abs"] / n["long"] if n["long"] else 0
    print(f"{str(r['month']):>10}  {r['long_to_nav']:8.2f}   {r['short_to_nav']:8.2f}   {sl_ref:5.3f}    {n['long_to_nav']:8.2f}   {n['short_to_nav']:8.2f}   {sl_new:5.3f}")
