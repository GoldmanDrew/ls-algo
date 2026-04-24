"""Tests for ``portfolio_ledger_dc`` / ``pnl_net_of_borrow_usd`` workbook semantics."""
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from scripts.daily_pair_workbook_formulas import (
    PORTFOLIO_LEDGER_DC_HEADERS,
    apply_portfolio_level_cost_model,
    compute_portfolio_ledger_dc_snapshot,
    reallocate_net_underlying_borrow_by_under,
)
from scripts.export_diamond_creek_daily_pair_workbook import export_diamond_creek_daily_pair_workbook
from scripts.export_dc_etf_arb_pairwise_workbook import _eom_v15_monthly_books, _lp_fee_row_formulas
from scripts.lp_fees_v15 import build_lp_fee_daily_cashflow_usd


def _minimal_bt(dates: list[pd.Timestamp]) -> pd.DataFrame:
    rows = []
    nav = 10_000_000.0
    cum = 0.0
    for d in dates:
        rows.append(
            {
                "nav": nav,
                "cum_costs": cum,
                "daily_borrow": 0.0,
                "daily_long_pnl": 0.0,
                "daily_short_pnl": 0.0,
                "cum_margin_debit": 0.0,
                "cum_margin_credit": 0.0,
            }
        )
        nav += 1000.0
        cum += 10.0
    return pd.DataFrame(rows, index=pd.DatetimeIndex(dates, name="date"))


class TestPortfolioLedgerDc(unittest.TestCase):
    def test_portfolio_level_cost_model_and_snapshot(self) -> None:
        d1 = pd.Timestamp("2023-01-03")
        d2 = pd.Timestamp("2023-01-04")
        ap = pd.DataFrame(
            {
                "date": [d1, d1, d2],
                "etf": ["E1", "E2", "E1"],
                "under": ["U1", "U2", "U1"],
                "long_sh": [10, 20, 10],
                "short_sh": [-5, -10, -5],
                "underlying_price": [100.0, 50.0, 101.0],
                "etf_price": [20.0, 10.0, 20.0],
                "long_margin_basis_usd": [1000.0, 1000.0, 1010.0],
                "short_financing_basis_usd": [0.0, 0.0, 0.0],
                "long_notional_usd": [1000.0, 1000.0, 1010.0],
                "short_notional_usd": [-100.0, -100.0, -100.0],
                "gross_notional_usd": [1100.0, 1100.0, 1110.0],
                "net_notional_usd": [900.0, 900.0, 910.0],
                "borrow_rate_annual": [0.05, 0.05, 0.05],
                "daily_long_pnl_usd": [10.0, 20.0, 5.0],
                "daily_short_pnl_usd": [-1.0, -2.0, -3.0],
                "daily_borrow_cost_usd": [1.0, 2.0, 4.0],
                "daily_underlying_borrow_cost_usd": [0.0, 0.0, 0.0],
                "daily_margin_debit_cost_usd": [1.0, 2.0, 3.0],
                "daily_short_credit_income_usd": [0.5, 0.5, 0.5],
                "daily_net_financing_cost_usd": [0.0, 0.0, 0.0],
                "daily_txn_cost_usd": [3.0, 4.0, 1.0],
                "daily_turnover_usd": [0.0, 0.0, 0.0],
                "daily_pair_gross_trading_pnl_usd": [9.0, 18.0, 2.0],
                "daily_pair_net_pnl_usd": [5.0, 14.0, -3.0],
                "daily_pair_net_ex_txn_usd": [8.0, 18.0, -2.0],
                "fed_funds_rate": [0.05, 0.05, 0.05],
                "is_rebal": [0, 0, 0],
            }
        )
        ap = apply_portfolio_level_cost_model(ap, verify_identities=True)
        self.assertTrue((ap["daily_txn_cost_usd"] == 0.0).all())
        self.assertTrue((ap["daily_margin_debit_cost_usd"] == 0.0).all())
        self.assertTrue((ap["daily_short_credit_income_usd"] == 0.0).all())
        got = compute_portfolio_ledger_dc_snapshot(
            ap,
            debit_margin_rate_annual=0.0432,
            financing_daycount=360.0,
            portfolio_daily_txn_by_date={d1: 7.0, d2: 1.0},
        )
        self.assertEqual(list(got.columns), list(PORTFOLIO_LEDGER_DC_HEADERS))
        self.assertEqual(len(got), 2)
        # Day 1: long notional 2000, txn sourced from portfolio_daily_txn_by_date.
        self.assertAlmostEqual(float(got.loc[0, "Long Notional USD"]), 2000.0)
        self.assertAlmostEqual(float(got.loc[0, "Margin Cost Book Level"]), 0.0)
        self.assertAlmostEqual(float(got.loc[0, "Daily T-Cost"]), 7.0)
        # pnl_net per row d1: (10-1-1)=8, (20-2-3)=15 -> need cum_borrow
        # cum row1 pair1: borrow 1 -> (10+(-1)-1)=8; pair2 borrow 2 cum 2 -> (20-2-2)=16? cum is per pair
        from scripts.daily_pair_workbook_formulas import _ensure_cum_borrow_cost_usd

        ap2 = _ensure_cum_borrow_cost_usd(ap)
        ap2["_pnl"] = (
            pd.to_numeric(ap2["daily_long_pnl_usd"], errors="coerce")
            + pd.to_numeric(ap2["daily_short_pnl_usd"], errors="coerce")
            - pd.to_numeric(ap2["cum_borrow_cost_usd"], errors="coerce")
        )
        g1 = float(ap2.loc[ap2["date"] == d1, "_pnl"].sum())
        self.assertAlmostEqual(float(got.loc[0, "Cumulative PnL Pre-T-Cost and Margin"]), g1)

    def test_export_contains_dc_formulas(self) -> None:
        d1 = pd.Timestamp("2023-01-03")
        d2 = pd.Timestamp("2023-01-04")
        ap = pd.DataFrame(
            {
                "date": [d1, d1, d2, d2],
                "etf": ["E1", "E2", "E1", "E2"],
                "under": ["U1", "U2", "U1", "U2"],
                "long_sh": [10, 20, 10, 20],
                "short_sh": [-5, -10, -5, -10],
                "underlying_price": [100.0, 50.0, 100.0, 50.0],
                "etf_price": [20.0, 10.0, 20.0, 10.0],
                "long_margin_basis_usd": [1000.0, 1000.0, 1000.0, 1000.0],
                "short_financing_basis_usd": [0.0, 0.0, 0.0, 0.0],
                "long_notional_usd": [1000.0, 1000.0, 1000.0, 1000.0],
                "short_notional_usd": [-100.0, -100.0, -100.0, -100.0],
                "gross_notional_usd": [1100.0, 1100.0, 1100.0, 1100.0],
                "net_notional_usd": [900.0, 900.0, 900.0, 900.0],
                "borrow_rate_annual": [0.05, 0.05, 0.05, 0.05],
                "daily_long_pnl_usd": [1.0, 2.0, 3.0, 4.0],
                "daily_short_pnl_usd": [-0.5, -1.0, -0.5, -1.0],
                "daily_borrow_cost_usd": [0.1, 0.2, 0.1, 0.2],
                "daily_underlying_borrow_cost_usd": [0.0, 0.0, 0.0, 0.0],
                "daily_margin_debit_cost_usd": [2.0, 2.0, 2.0, 2.0],
                "daily_short_credit_income_usd": [1.0, 1.0, 1.0, 1.0],
                "daily_net_financing_cost_usd": [0.0, 0.0, 0.0, 0.0],
                "daily_txn_cost_usd": [0.5, 0.5, 0.25, 0.25],
                "daily_turnover_usd": [0.0, 0.0, 0.0, 0.0],
                "daily_pair_gross_trading_pnl_usd": [0.5, 1.0, 2.5, 3.0],
                "daily_pair_net_pnl_usd": [0.4, 0.8, 2.25, 2.7],
                "daily_pair_net_ex_txn_usd": [0.9, 1.3, 2.5, 2.95],
                "fed_funds_rate": [0.05, 0.05, 0.05, 0.05],
                "is_rebal": [0, 0, 0, 0],
            }
        )
        bt = _minimal_bt([d1, d2])
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            export_diamond_creek_daily_pair_workbook(
                ap,
                bt,
                p,
                full_date_range=True,
                include_only_active_pairs=False,
                include_per_pair_sheets=False,
                use_portfolio_excel_formulas=True,
                fund_attribution_xlsx=None,
                debit_margin_rate_annual=0.05,
                include_portfolio_ledger_dc=True,
            )
            wb = load_workbook(p, data_only=False)
            self.assertIn("portfolio_ledger_dc", wb.sheetnames)
            self.assertIn("dc_workbook_settings", wb.sheetnames)
            ws = wb["portfolio_ledger_dc"]
            b2 = ws["B2"].value
            self.assertIsInstance(b2, str)
            self.assertIn("SUMIFS", b2)
            self.assertIn("ALL_PAIRS", b2)
            e2 = ws["E2"].value
            self.assertIsInstance(e2, str)
            self.assertIn("VLOOKUP", str(e2))
            self.assertIn("book_daily", str(e2))
            g2 = ws["G2"].value
            self.assertIsInstance(g2, str)
            self.assertIn("SUMIFS", str(g2))
            h_ap = {str(c.value): c.column for c in wb["ALL_PAIRS"][1] if c.value}
            self.assertIn("pnl_net_of_borrow_usd", h_ap)
            ad_letter = get_column_letter(h_ap["pnl_net_of_borrow_usd"])
            self.assertIn(ad_letter, str(g2))
            # Portfolio-level model: pair txn/margin/short-credit are hard zeroed.
            hdr = {str(c.value): c.column for c in wb["ALL_PAIRS"][1] if c.value}
            self.assertEqual(wb["ALL_PAIRS"].cell(2, hdr["daily_txn_cost_usd"]).value, 0)
            self.assertEqual(wb["ALL_PAIRS"].cell(2, hdr["daily_margin_debit_cost_usd"]).value, 0)
            self.assertEqual(wb["ALL_PAIRS"].cell(2, hdr["daily_short_credit_income_usd"]).value, 0)
            wb.close()

    def test_net_underlying_borrow_respects_offsetting_long_sh(self) -> None:
        d = pd.Timestamp("2023-01-03")
        ap = pd.DataFrame(
            {
                "date": [d, d],
                "etf": ["TSLQ", "TSLR"],
                "under": ["TSLA", "TSLA"],
                "long_sh": [-10.0, 10.0],
                "short_sh": [0.0, 0.0],
                "underlying_price": [200.0, 200.0],
                "etf_price": [50.0, 50.0],
                "borrow_rate_annual": [0.0, 0.0],
                "under_borrow_rate_annual": [0.1, 0.1],
                "daily_underlying_borrow_cost_usd": [999.0, 999.0],
                "daily_borrow_cost_usd": [0.0, 0.0],
            }
        )
        out = reallocate_net_underlying_borrow_by_under(ap, trading_days=252.0)
        self.assertTrue((out["daily_underlying_borrow_cost_usd"] == 0.0).all())
        out2 = reallocate_net_underlying_borrow_by_under(
            ap.assign(long_sh=[-10.0, 0.0], short_sh=[-1.0, 0.0], borrow_rate_annual=[0.05, 0.0], etf_price=[10.0, 10.0]),
            trading_days=252.0,
        )
        s = out2["daily_underlying_borrow_cost_usd"].sum()
        self.assertGreater(s, 0.0)
        # Single net short row gets full accrual for 10 sh short on underlying.
        p, br, td = 200.0, 0.1, 252.0
        expect = 10.0 * p * br / td
        self.assertAlmostEqual(float(s), expect, places=4)


class TestLpFeeDailyCashflow(unittest.TestCase):
    def test_quarterly_mgmt_hits_on_last_month_for_2023(self) -> None:
        d = pd.bdate_range("2023-01-03", "2023-03-31", freq="B")
        nav = pd.Series(10_000_000.0, index=d)
        out = build_lp_fee_daily_cashflow_usd(
            nav,
            d,
            attribution_base_capital=10_000_000.0,
            management_fee_annual=0.04,
            incentive_fee=0.0,
        )
        q1 = d[d.to_period("Q") == pd.Period("2023Q1", "Q")]
        first, last = q1[0], q1[-1]
        self.assertEqual(float(out.loc[first, "mgmt_usd"]), 0.0)
        self.assertGreater(float(out.loc[last, "mgmt_usd"]), 0.0)

    def test_quarterly_mgmt_hits_on_last_trading_day_pre_2023(self) -> None:
        d = pd.bdate_range("2022-10-03", "2022-12-30", freq="B")
        nav = pd.Series(10_000_000.0, index=d)
        out = build_lp_fee_daily_cashflow_usd(
            nav,
            d,
            attribution_base_capital=10_000_000.0,
            management_fee_annual=0.04,
            incentive_fee=0.0,
        )
        q4 = d[d.to_period("Q") == pd.Period("2022Q4", "Q")]
        first, last = q4[0], q4[-1]
        self.assertEqual(float(out.loc[first, "mgmt_usd"]), 0.0)
        self.assertGreater(float(out.loc[last, "mgmt_usd"]), 0.0)

    def test_at_most_one_management_fee_per_calendar_quarter(self) -> None:
        d = pd.bdate_range("2022-10-04", "2023-12-29", freq="B")
        nav = pd.Series(1e7, index=d) + 0.0
        out = build_lp_fee_daily_cashflow_usd(
            nav, d, attribution_base_capital=1e7, management_fee_annual=0.04, incentive_fee=0.0
        )
        counts = out["mgmt_usd"].gt(0).groupby(d.to_period("Q")).sum()
        self.assertTrue(
            (counts <= 1).all(),
            msg="at most one day with positive mgmt per quarter",
        )

    def test_q2_2023_mgmt_appears_in_june(self) -> None:
        d = pd.bdate_range("2023-01-02", "2023-12-29", freq="B")
        nav = pd.Series(1e7, index=d) + 0.0
        out = build_lp_fee_daily_cashflow_usd(
            nav, d, attribution_base_capital=1e7, management_fee_annual=0.04, incentive_fee=0.0
        )
        q2 = d[(d.to_period("Q") == pd.Period("2023Q2", "Q"))]
        q2_last = q2[-1]
        in_june = q2_last.month == 6
        self.assertTrue(
            in_june, msg="last Q2 2023 trading day should be in final month (June) of the quarter"
        )
        self.assertGreater(float(out.loc[q2_last, "mgmt_usd"]), 0.0)


class TestLpFeeExcelFormulas(unittest.TestCase):
    def test_lp_fee_row_strings_reference_sheets(self) -> None:
        f = _lp_fee_row_formulas(2, daily_pnl_b_row=3)
        self.assertIn("Daily PnL", f["A"])
        self.assertIn("book_raw", f["D"])
        self.assertIn("MAXIFS", f["F"])
        self.assertIn("dc_pairwise_params!$B$3/4", f["B"])


class TestEomV15Books(unittest.TestCase):
    def test_lmb_sfb_on_end_of_sim_day(self) -> None:
        d0 = pd.Timestamp("2023-01-10")
        pdw = pd.DataFrame(
            {
                "date": [d0, d0],
                "long_sh": [10.0, 0.0],
                "short_sh": [0.0, -2.0],
                "underlying_price": [200.0, 200.0],
                "etf_price": [50.0, 50.0],
            }
        )
        pdw["long_notional_usd"] = pdw["long_sh"] * pdw["underlying_price"]
        pdw["short_notional_usd"] = pdw["short_sh"] * pdw["etf_price"]
        lsh, ssh = pdw["long_sh"], pdw["short_sh"]
        pu, pe = pdw["underlying_price"], pdw["etf_price"]
        pdw["long_margin_basis_usd"] = np.maximum(lsh, 0.0) * pu
        pdw["short_financing_basis_usd"] = np.maximum(-lsh, 0.0) * pu + np.maximum(-ssh, 0.0) * pe
        a, b = _eom_v15_monthly_books(pdw, d0)
        self.assertAlmostEqual(a, 2_000.0)  # 10*200
        self.assertAlmostEqual(b, 0.0 + 2.0 * 50.0)  # -ssh short etf: 2*50


if __name__ == "__main__":
    unittest.main()
