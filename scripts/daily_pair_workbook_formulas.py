"""
Inject Excel **formulas** into the Daily Pair workbook so notionals / margin /
financing / net PnL identities match ``build_reconciliation_daily`` semantics.

**Notional and book columns (v15, shared with DC ETF pairwise export & Monthly F/G)**

* **Signed leg notionals:** ``long_notional_usd = long_sh * underlying_price``;
  ``short_notional_usd = short_sh * etf_price``. Use for *signed* leg mark value.
* **Gross size (all legs, unsigned):** ``gross_notional_usd = ABS(long_sh)*PU+ABS(ssh)*ETF``.
* **“Long book” / “Short financing book” (margin & borrow, not the same as signed
  leg notionals):** ``long_margin_basis_usd = MAX(long_sh,0)*PU``;
  ``short_financing_basis_usd = MAX(-long_sh,0)*PU+MAX(-short_sh,0)*ETF`` — these
  are what ``export_dc_etf_arb_pairwise_workbook`` uses for Monthly F (sum LMB)
  and G (sum SFB) on EOM.

**Excel borrow vs. ``reallocate_net_underlying_borrow_by_under``:** the inject
formulas in ``_inject_all_pairs_leg_pnl_borrow_excel`` use *per-row* accrual.
:func:`reallocate_net_underlying_borrow_by_under` recomputes underlying-borrow
by net short underlying per (date, under) in *Python* and can change
``daily_borrow`` totals; rollups may not match a naive ``SUM`` of the Excel
``ALL_PAIRS`` borrow column when the reallocator runs.

DC reference workbook mapping (``Diamond_Creek_Daily_Pairs.xlsx`` vs v15 export)
===============================================================================
We keep **v15** ``ALL_PAIRS`` column order (``_all_pairs_columns``) and the
existing ``portfolio_ledger`` engine sheet (Option **B**: maintainability).
A second sheet **``portfolio_ledger_dc``** clones the reference **nine-column**
portfolio layout and formula *patterns*, but every ``SUMIFS`` targets the real
v15 header names / column letters on ``ALL_PAIRS``.

Reference ``ALL_PAIRS`` (legacy) → v15 (this repo)
  - ``long_notional_usd`` (ref col **H**) → ``long_notional_usd`` (v15 col **J**)
  - ``short_notional_usd`` (ref **I**) → ``short_notional_usd`` (v15 **K**)
  - ``gross_notional_usd`` (ref **J**, ``=H-I``) → ``gross_notional_usd`` (v15 **L**,
    formula ``ABS(long_sh)*underlying + ABS(short_sh)*etf`` — not identical to
    ``H-I``; per-pair economics follow the engine.)
  - ``daily_borrow_cost_usd`` (ref **L**, ``-I*K/252``) → v15 **Q** (engine value;
    day-count differs from the legacy 252 short-leg accrual.)
  - ``PnL Net of Borrow`` (ref **X**, ``=W-N`` with per-pair leg links) →
    **``pnl_net_of_borrow_usd``** (v15 trailing col **AD**): ``=daily_long_pnl_usd
    + daily_short_pnl_usd - cum_borrow_cost_usd`` where ``cum_borrow_cost_usd``
    (**AC**) is exported as values (running sum of **Q** per pair). This matches
    the reference *intent* (leg price P&L minus cumulative borrow) without
    fragile cross-sheet ``R$2`` links.

Reference ``portfolio_ledger`` → ``portfolio_ledger_dc`` (this repo)
  - **B** ``Long Notional USD`` → ``SUMIFS`` on v15 ``long_notional_usd`` (**J**),
    *not* ``long_margin_basis_usd`` (**H**); the legacy engine ``portfolio_ledger``
    still sums **H** for portfolio margin — intentionally unchanged.
  - **C** ``Debit Margin Rate`` → ``dc_workbook_settings!$B$2`` (default **0.0432**),
    replacing the reference literal ``=0.0432`` while staying cell-driven.
  - **D** ``Margin Cost Book Level`` → ``IF(ROW()=2,0,B*C/daycount)`` with
    ``daycount`` from ``dc_workbook_settings!$B$1`` (default **360**).
  - **E** ``Daily T-Cost`` → ``SUMIFS`` on ``daily_txn_cost_usd`` (**V**) by date
    (pair txn roll-up; the reference sometimes used book-level T-cost in **E** —
    we document that difference; acceptance tests use **ALL_PAIRS** inputs only).
  - **G** (reference name: cumulative block sum) → ``SUMIFS`` on **AD** by date,
    replacing ``SUM(OFFSET(…, 26 rows))`` so variable pairs/day work.
  - **H** ``=G-F``, **I** first-row ``=H2`` then ``=H{n}-H{n-1}`` as in the reference.

**Value-only (explicit marks / static inputs):** ``date``, ``etf``, ``under``,
``long_sh``, ``short_sh``, ``underlying_price``, ``etf_price``,
``borrow_rate_annual``, ``under_borrow_rate_annual``, ``fed_funds_rate``,
``is_rebal``,
and Excel-style **cumulative trade-cash** inputs ``excel_long_trades_usd`` /
``excel_short_trades_usd`` (engine output).

When those trade-cash columns are present, ``inject_daily_pair_workbook_formulas`` also
rewrites one-row **v15** leg P&L and borrow to Excel formulas
(``=SUM(long_level)`` diffs, ``ABS(short) * mark * rate / VLOOKUP(trading_days)``) so
``daily_{long,short,borrow,underlying}_*`` are no longer ``openpyxl``-written numbers
before the algebraic (notional / net) post-pass.

**Value-only (still engine in Python when trade-cash columns are missing):**
``daily_long_pnl_usd``, ``daily_short_pnl_usd``,
``daily_borrow_cost_usd``, ``daily_underlying_borrow_cost_usd`` (fallback),
``daily_margin_debit_cost_usd``, ``daily_short_credit_income_usd``,
``daily_txn_cost_usd``, ``daily_turnover_usd``,

**Formula-driven (algebraic identities on the same row):**
``long_margin_basis_usd``, ``short_financing_basis_usd``, ``long_notional_usd``,
``short_notional_usd``, ``gross_notional_usd``, ``net_notional_usd``,
``daily_net_financing_cost_usd`` (= margin debit − short credit),
``daily_pair_gross_trading_pnl_usd`` (= long leg PnL + short leg PnL),
``daily_pair_net_pnl_usd`` (= gross trading − borrow − net financing − txn),
``daily_pair_net_ex_txn_usd`` (= net + txn).

Portfolio ``net_after_portfolio_margin_and_pair_txn_usd`` is
``SUM(daily_pair_net_pnl_usd) − daily_portfolio_margin_usd`` only — pair txn is
already inside ``daily_pair_net_pnl_usd`` and must not be subtracted again.

``portfolio_ledger`` uses ``SUMIFS`` / ``AVERAGEIFS`` into ``ALL_PAIRS`` and
``VLOOKUP`` into ``book_daily`` for book NAV and book daily txn (still sourced
from the backtest export; txn as first-row-aware diff is written to
``book_daily`` to avoid fragile OFFSET chains on row 1).
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook


SKIP_SHEETS = frozenset(
    {
        "book_daily",
        "checks",
        "dc_workbook_settings",
        "portfolio_ledger_dc",
    }
)

# ``portfolio_ledger`` column order (must match ``export_diamond_creek_daily_pair_workbook``).
PORTFOLIO_LEDGER_COLUMNS: tuple[str, ...] = (
    "date",
    "total_long_notional_usd",
    "benchmark_rate_annual",
    "margin_spread_annual",
    "effective_margin_rate_annual",
    "daily_portfolio_margin_usd",
    "daily_txn_pairs_sum_usd",
    "sum_pair_long_pnl_usd",
    "sum_pair_short_pnl_usd",
    "sum_pair_gross_trading_pnl_usd",
    "book_daily_txn_cost_usd",
    "book_daily_net_margin_usd",
    "book_txn_plus_net_margin_usd",
    "sum_pair_daily_net_pnl_usd",
    "net_after_portfolio_margin_and_pair_txn_usd",
    "cum_gross_trading_pnl_usd",
    "cum_pair_net_pnl_usd",
    "book_daily_nav_change_usd",
    "daily_pnl_book_usd",
    "fund_daily_pnl_usd",
    "diff_model_net_minus_fund_usd",
)

# ``portfolio_ledger_dc`` mirrors ``Diamond_Creek_Daily_Pairs.xlsx`` ``portfolio_ledger``.
PORTFOLIO_LEDGER_DC_HEADERS: tuple[str, ...] = (
    "Date",
    "Long Notional USD",
    "Debit Margin Rate",
    "Margin Cost Book Level",
    "Daily T-Cost",
    "Margin + T-Cost",
    "Cumulative PnL Pre-T-Cost and Margin",
    "Cumulative Pre-Fee PnL",
    "Daily PnL",
)


def fill_portfolio_ledger_values(
    all_pairs: pd.DataFrame,
    book_daily: pd.DataFrame,
    *,
    margin_spread_annual: float,
    financing_daycount: float = 360.0,
    fund_by_date: dict[pd.Timestamp, float] | None = None,
) -> pd.DataFrame:
    """Populate ``portfolio_ledger`` with numeric values (no Excel formulas)."""
    if all_pairs.empty:
        return pd.DataFrame(columns=list(PORTFOLIO_LEDGER_COLUMNS))
    ap = apply_portfolio_level_cost_model(all_pairs, verify_identities=True)
    g = ap.groupby("date", as_index=False).agg(
        total_long_notional_usd=("long_margin_basis_usd", "sum"),
        benchmark_rate_annual=("fed_funds_rate", "mean"),
        sum_pair_long_pnl_usd=("daily_long_pnl_usd", "sum"),
        sum_pair_short_pnl_usd=("daily_short_pnl_usd", "sum"),
        sum_pair_daily_net_pnl_usd=("daily_pair_net_pnl_usd", "sum"),
    )
    g["margin_spread_annual"] = float(margin_spread_annual)
    g["effective_margin_rate_annual"] = g["benchmark_rate_annual"] + g["margin_spread_annual"]
    g["daily_portfolio_margin_usd"] = (
        g["total_long_notional_usd"] * g["effective_margin_rate_annual"] / float(financing_daycount)
    )
    g["sum_pair_gross_trading_pnl_usd"] = g["sum_pair_long_pnl_usd"] + g["sum_pair_short_pnl_usd"]
    bd = book_daily.copy()
    if bd.empty:
        m = g.copy()
        m["book_daily_txn_cost_usd"] = 0.0
        m["book_daily_net_margin_usd"] = 0.0
        m["book_daily_nav_change_usd"] = np.nan
        m["daily_txn_pairs_sum_usd"] = 0.0
    else:
        bd["date"] = pd.to_datetime(bd["date"]).dt.normalize()
        m = g.merge(bd, on="date", how="left")
        m["book_daily_txn_cost_usd"] = pd.to_numeric(m["book_daily_txn_usd"], errors="coerce").fillna(0.0)
        m["book_daily_net_margin_usd"] = pd.to_numeric(m["book_daily_net_margin_usd"], errors="coerce").fillna(0.0)
        m["book_daily_nav_change_usd"] = pd.to_numeric(m["book_nav_change_usd"], errors="coerce")
        # Portfolio transaction cost is charged once at book level.
        m["daily_txn_pairs_sum_usd"] = pd.to_numeric(m["book_daily_txn_cost_usd"], errors="coerce").fillna(0.0)
    m["book_txn_plus_net_margin_usd"] = (
        pd.to_numeric(m["book_daily_txn_cost_usd"], errors="coerce").fillna(0.0)
        + pd.to_numeric(m["book_daily_net_margin_usd"], errors="coerce").fillna(0.0)
    )
    # Portfolio txn/margin are charged once at portfolio level.
    m["net_after_portfolio_margin_and_pair_txn_usd"] = (
        pd.to_numeric(m["sum_pair_daily_net_pnl_usd"], errors="coerce").fillna(0.0)
        - pd.to_numeric(m["daily_portfolio_margin_usd"], errors="coerce").fillna(0.0)
        - pd.to_numeric(m["daily_txn_pairs_sum_usd"], errors="coerce").fillna(0.0)
    )
    m["daily_pnl_book_usd"] = m["book_daily_nav_change_usd"]
    if fund_by_date:
        m["fund_daily_pnl_usd"] = m["date"].map(lambda d: fund_by_date.get(pd.Timestamp(d).normalize(), np.nan))
    else:
        m["fund_daily_pnl_usd"] = np.nan
    m["diff_model_net_minus_fund_usd"] = m["net_after_portfolio_margin_and_pair_txn_usd"] - pd.to_numeric(
        m["fund_daily_pnl_usd"], errors="coerce"
    )
    m["cum_gross_trading_pnl_usd"] = pd.to_numeric(m["sum_pair_gross_trading_pnl_usd"], errors="coerce").cumsum()
    m["cum_pair_net_pnl_usd"] = pd.to_numeric(m["sum_pair_daily_net_pnl_usd"], errors="coerce").cumsum()
    out = m[list(PORTFOLIO_LEDGER_COLUMNS)].copy()
    return out


def reallocate_net_underlying_borrow_by_under(
    all_pairs: pd.DataFrame,
    *,
    trading_days: float = 252.0,
) -> pd.DataFrame:
    """
    Reallocate ``daily_underlying_borrow_cost_usd`` by ``(date, under)`` so
    accrual uses **net** short underlying exposure (``long_sh``), not per row.

    The naive per-row accrual ``IF(long_sh<0, |long_sh|*P*R/τ,0)`` double-counts
    when several pair rows (e.g. same underlying in bucket-3 style structures)
    have offsetting long_sh.  Borrow is charged only on ``-min(0, sum(long_sh))``,
    split to rows in proportion to ``-min(0, long_sh_i)``.

    Rebuilds ``daily_borrow_cost_usd`` = ETF short-leg accrual + the new
    underlying slice; call **before** :func:`apply_portfolio_level_cost_model`.

    **Attribution workbooks** that re-inject per-row short×rate *Excel* formulas
    (``_inject_all_pairs_leg_pnl_borrow_excel``) will show different daily borrow
    *sums* than a ``SUM`` of the rows after this step when offsetting
    long-shares exist on the same underlying. Python rollups use the reallocated
    series; see ``export_dc_etf_arb_pairwise_workbook`` module docstring.
    """
    if all_pairs.empty:
        return all_pairs.copy()
    out = all_pairs.copy()
    out["date"] = pd.to_datetime(out["date"], errors="coerce").dt.normalize()
    td = float(trading_days) if trading_days and float(trading_days) > 0 else 252.0
    for c in (
        "long_sh",
        "short_sh",
        "underlying_price",
        "etf_price",
        "borrow_rate_annual",
        "under_borrow_rate_annual",
    ):
        if c not in out.columns:
            out[c] = np.nan
    bru = pd.to_numeric(out.get("under_borrow_rate_annual"), errors="coerce")
    if bru.isna().all() and "borrow_rate_annual" in out.columns:
        bru = pd.to_numeric(out["borrow_rate_annual"], errors="coerce")
    out["under_borrow_rate_annual"] = bru.fillna(0.0)
    out["borrow_rate_annual"] = pd.to_numeric(out.get("borrow_rate_annual"), errors="coerce").fillna(0.0)
    out["long_sh"] = pd.to_numeric(out.get("long_sh"), errors="coerce").fillna(0.0)
    out["short_sh"] = pd.to_numeric(out.get("short_sh"), errors="coerce").fillna(0.0)
    out["underlying_price"] = pd.to_numeric(out.get("underlying_price"), errors="coerce").fillna(0.0)
    out["etf_price"] = pd.to_numeric(out.get("etf_price"), errors="coerce").fillna(0.0)
    if "daily_underlying_borrow_cost_usd" not in out.columns:
        out["daily_underlying_borrow_cost_usd"] = 0.0
    if "daily_borrow_cost_usd" not in out.columns:
        out["daily_borrow_cost_usd"] = 0.0

    parts: list[pd.DataFrame] = []
    for _, g in out.groupby(["date", "under"], sort=True):
        s = g.copy()
        lshv = s["long_sh"].to_numpy(dtype=float, copy=False)
        net_l = float(lshv.sum())
        sh_mass = float((-np.minimum(lshv, 0.0)).sum())
        p_mean = float(s["underlying_price"].mean()) if not s["underlying_price"].isna().all() else 0.0
        br_m = float(s["under_borrow_rate_annual"].mean())
        u_total = 0.0
        if net_l < 0.0 and sh_mass > 1e-12:
            u_total = max(0.0, -net_l) * p_mean * br_m / td
        w = np.zeros(len(s), dtype=float)
        if u_total > 0.0 and sh_mass > 1e-12:
            w = (-np.minimum(lshv, 0.0)) / sh_mass
        s = s.copy()
        s["daily_underlying_borrow_cost_usd"] = u_total * w
        s_shv = s["short_sh"].to_numpy(dtype=float, copy=False)
        pev = s["etf_price"].to_numpy(dtype=float, copy=False)
        brev = s["borrow_rate_annual"].to_numpy(dtype=float, copy=False)
        b_etf = np.where(s_shv < 0, np.abs(s_shv) * pev * brev / td, 0.0)
        s["daily_borrow_cost_usd"] = b_etf + s["daily_underlying_borrow_cost_usd"].to_numpy(dtype=float, copy=False)
        parts.append(s)
    merged = pd.concat(parts, ignore_index=True)
    return merged.sort_values(["date", "under", "etf"], kind="mergesort").reset_index(drop=True)


def apply_portfolio_level_cost_model(
    all_pairs: pd.DataFrame,
    *,
    verify_identities: bool = True,
    tol_usd: float = 1e-6,
) -> pd.DataFrame:
    """
    Enforce portfolio-level transaction/margin accounting on pair rows.

    Recomputes signed leg notionals, **LMB** / **SFB**, and gross/net notional
    from shares and prices (same as ``_inject_all_pairs_formulas``; see module
    docstring).

    Pair rows keep leg P&L and borrow costs. Pair txn, pair margin debit, and
    short-credit income are set to zero; pair net is recomputed as:

    ``long + short - borrow - underlying_borrow``.
    """
    if all_pairs.empty:
        return all_pairs.copy()
    out = all_pairs.copy()
    req = (
        "long_sh",
        "short_sh",
        "underlying_price",
        "etf_price",
        "daily_long_pnl_usd",
        "daily_short_pnl_usd",
        "daily_borrow_cost_usd",
        "daily_underlying_borrow_cost_usd",
    )
    for c in req:
        if c not in out.columns:
            out[c] = 0.0
    for c in (
        "daily_margin_debit_cost_usd",
        "daily_short_credit_income_usd",
        "daily_txn_cost_usd",
        "daily_net_financing_cost_usd",
    ):
        if c not in out.columns:
            out[c] = 0.0
    lsh = pd.to_numeric(out["long_sh"], errors="coerce").fillna(0.0)
    ssh = pd.to_numeric(out["short_sh"], errors="coerce").fillna(0.0)
    pu = pd.to_numeric(out["underlying_price"], errors="coerce").fillna(0.0)
    pe = pd.to_numeric(out["etf_price"], errors="coerce").fillna(0.0)
    out["long_notional_usd"] = lsh * pu
    out["short_notional_usd"] = ssh * pe
    out["long_margin_basis_usd"] = np.maximum(lsh, 0.0) * pu
    out["short_financing_basis_usd"] = np.maximum(-lsh, 0.0) * pu + np.maximum(-ssh, 0.0) * pe
    out["gross_notional_usd"] = np.abs(lsh) * pu + np.abs(ssh) * pe
    out["net_notional_usd"] = out["long_notional_usd"] + out["short_notional_usd"]
    lp = pd.to_numeric(out["daily_long_pnl_usd"], errors="coerce").fillna(0.0)
    sp = pd.to_numeric(out["daily_short_pnl_usd"], errors="coerce").fillna(0.0)
    bor = pd.to_numeric(out["daily_borrow_cost_usd"], errors="coerce").fillna(0.0)
    ubor = pd.to_numeric(out["daily_underlying_borrow_cost_usd"], errors="coerce").fillna(0.0)
    out["daily_margin_debit_cost_usd"] = 0.0
    out["daily_short_credit_income_usd"] = 0.0
    out["daily_txn_cost_usd"] = 0.0
    out["daily_net_financing_cost_usd"] = 0.0
    out["daily_pair_gross_trading_pnl_usd"] = lp + sp
    out["daily_pair_net_pnl_usd"] = out["daily_pair_gross_trading_pnl_usd"] - bor - ubor
    out["daily_pair_net_ex_txn_usd"] = out["daily_pair_net_pnl_usd"]
    if verify_identities:
        if float(np.abs(pd.to_numeric(out["daily_short_credit_income_usd"], errors="coerce").fillna(0.0)).max()) > tol_usd:
            raise ValueError("daily_short_credit_income_usd must be zero under portfolio-level model")
        if float(np.abs(pd.to_numeric(out["daily_margin_debit_cost_usd"], errors="coerce").fillna(0.0)).max()) > tol_usd:
            raise ValueError("daily_margin_debit_cost_usd must be zero under portfolio-level model")
        if float(np.abs(pd.to_numeric(out["daily_txn_cost_usd"], errors="coerce").fillna(0.0)).max()) > tol_usd:
            raise ValueError("daily_txn_cost_usd must be zero under portfolio-level model")
    return out


def _ensure_cum_borrow_cost_usd(all_pairs: pd.DataFrame) -> pd.DataFrame:
    """Per (under, etf) cumulative sum of ``daily_borrow_cost_usd`` by ``date``."""
    if all_pairs.empty:
        return all_pairs
    out = all_pairs.sort_values(["date", "under", "etf"], kind="mergesort").copy()
    b = pd.to_numeric(out["daily_borrow_cost_usd"], errors="coerce").fillna(0.0)
    keys = [out["under"].astype(str), out["etf"].astype(str)]
    out["cum_borrow_cost_usd"] = b.groupby(keys, sort=False).cumsum()
    return out


def compute_portfolio_ledger_dc_snapshot(
    all_pairs: pd.DataFrame,
    *,
    debit_margin_rate_annual: float,
    financing_daycount: float = 360.0,
    portfolio_daily_txn_by_date: dict[pd.Timestamp, float] | None = None,
) -> pd.DataFrame:
    """
    Reproduce ``portfolio_ledger_dc`` numeric semantics in pandas (tests / gold).

    Uses only ``ALL_PAIRS``-level inputs: long notionals, txn, leg P&L, and
    ``cum_borrow_cost_usd`` (adds it via ``_ensure_cum_borrow_cost_usd`` if absent).
    """
    if all_pairs.empty:
        return pd.DataFrame(columns=list(PORTFOLIO_LEDGER_DC_HEADERS))
    ap = apply_portfolio_level_cost_model(all_pairs, verify_identities=True)
    ap = _ensure_cum_borrow_cost_usd(ap)
    ap["date"] = pd.to_datetime(ap["date"], errors="coerce").dt.normalize()
    ap["_pnl_net"] = (
        pd.to_numeric(ap["daily_long_pnl_usd"], errors="coerce").fillna(0.0)
        + pd.to_numeric(ap["daily_short_pnl_usd"], errors="coerce").fillna(0.0)
        - pd.to_numeric(ap["cum_borrow_cost_usd"], errors="coerce").fillna(0.0)
    )
    g = (
        ap.groupby("date", as_index=False)
        .agg(
            long_notional_sum=("long_notional_usd", "sum"),
            pnl_net_sum=("_pnl_net", "sum"),
        )
        .sort_values("date", kind="mergesort")
        .reset_index(drop=True)
    )
    g["Debit Margin Rate"] = float(debit_margin_rate_annual)
    dc = float(financing_daycount)
    g["Margin Cost Book Level"] = g["long_notional_sum"] * g["Debit Margin Rate"] / dc
    g.loc[g.index[0], "Margin Cost Book Level"] = 0.0
    if portfolio_daily_txn_by_date:
        g["Daily T-Cost"] = g["date"].map(
            lambda d: float(portfolio_daily_txn_by_date.get(pd.Timestamp(d).normalize(), 0.0))
        )
    else:
        g["Daily T-Cost"] = 0.0
    g["Margin + T-Cost"] = g["Margin Cost Book Level"] + g["Daily T-Cost"]
    g["Cumulative PnL Pre-T-Cost and Margin"] = g["pnl_net_sum"]
    g["Cumulative Pre-Fee PnL"] = g["Cumulative PnL Pre-T-Cost and Margin"] - g["Margin + T-Cost"]
    g["Daily PnL"] = g["Cumulative Pre-Fee PnL"].diff()
    g.loc[g.index[0], "Daily PnL"] = g.loc[g.index[0], "Cumulative Pre-Fee PnL"]
    out = pd.DataFrame(
        {
            "Date": g["date"],
            "Long Notional USD": g["long_notional_sum"],
            "Debit Margin Rate": g["Debit Margin Rate"],
            "Margin Cost Book Level": g["Margin Cost Book Level"],
            "Daily T-Cost": g["Daily T-Cost"],
            "Margin + T-Cost": g["Margin + T-Cost"],
            "Cumulative PnL Pre-T-Cost and Margin": g["Cumulative PnL Pre-T-Cost and Margin"],
            "Cumulative Pre-Fee PnL": g["Cumulative Pre-Fee PnL"],
            "Daily PnL": g["Daily PnL"],
        }
    )
    return out[list(PORTFOLIO_LEDGER_DC_HEADERS)].reset_index(drop=True)


def normalize_series_for_book_merge(s: pd.Series) -> pd.Series:
    """Align ``pair_daily['date']`` with ``ALL_BT`` index (naive midnight)."""
    out = pd.to_datetime(s, errors="coerce")
    try:
        if out.dt.tz is not None:
            out = out.dt.tz_convert("UTC").dt.tz_localize(None)
    except (TypeError, AttributeError, ValueError):
        pass
    return out.dt.normalize()


def _naive_normalized(ts: pd.Timestamp | np.datetime64 | str | float) -> pd.Timestamp:
    t = pd.Timestamp(ts)
    if t.tzinfo is not None:
        t = t.tz_convert("UTC").tz_localize(None)
    return t.normalize()


def build_book_daily_dataframe(
    bt: pd.DataFrame,
    dates: Iterable[pd.Timestamp | np.datetime64 | str],
    *,
    attribution_base_capital: float,
) -> pd.DataFrame:
    """
    One row per calendar date in ``dates`` that exists in ``bt`` (sorted).
    Columns match ``VLOOKUP`` targets on ``checks`` / ``portfolio_ledger``:
    ``date``, ``nav``, ``book_daily_txn_usd``, ``book_nav_change_usd``,
    ``book_daily_borrow_usd``, ``book_daily_long_short_pnl_usd``,
    ``book_daily_net_margin_usd``.
    """
    b = bt.sort_index().copy()
    b.index = pd.to_datetime(b.index)
    if b.index.tz is not None:
        b.index = b.index.tz_convert("UTC").tz_localize(None)
    b.index = b.index.normalize()
    ux = sorted({_naive_normalized(x) for x in dates})
    rows: list[dict[str, float | pd.Timestamp]] = []
    for d in ux:
        if d not in b.index:
            continue
        loc = b.index.get_loc(d)
        if isinstance(loc, slice):
            loc = int(loc.start)
        elif isinstance(loc, np.ndarray):
            loc = int(loc.flat[0])
        else:
            loc = int(loc)
        nav = float(pd.to_numeric(b.iloc[loc]["nav"], errors="coerce"))
        cum = float(pd.to_numeric(b.iloc[loc]["cum_costs"], errors="coerce"))
        brw = float(pd.to_numeric(b.iloc[loc].get("daily_borrow"), errors="coerce") or 0.0)
        ls = float(pd.to_numeric(b.iloc[loc].get("daily_long_pnl"), errors="coerce") or 0.0) + float(
            pd.to_numeric(b.iloc[loc].get("daily_short_pnl"), errors="coerce") or 0.0
        )
        prev_nav = float(attribution_base_capital) if loc == 0 else float(b.iloc[loc - 1]["nav"])
        prev_cum = 0.0 if loc == 0 else float(b.iloc[loc - 1]["cum_costs"])
        md = float(pd.to_numeric(b.iloc[loc].get("cum_margin_debit"), errors="coerce") or 0.0)
        mc = float(pd.to_numeric(b.iloc[loc].get("cum_margin_credit"), errors="coerce") or 0.0)
        pre_md = 0.0 if loc == 0 else float(pd.to_numeric(b.iloc[loc - 1].get("cum_margin_debit"), errors="coerce") or 0.0)
        pre_mc = 0.0 if loc == 0 else float(pd.to_numeric(b.iloc[loc - 1].get("cum_margin_credit"), errors="coerce") or 0.0)
        net_margin = (md - pre_md) - (mc - pre_mc)
        rows.append(
            {
                "date": d,
                "nav": nav,
                "book_daily_txn_usd": cum - prev_cum,
                "book_nav_change_usd": nav - prev_nav,
                "book_daily_borrow_usd": brw,
                "book_daily_long_short_pnl_usd": ls,
                "book_daily_net_margin_usd": net_margin,
            }
        )
    return pd.DataFrame(rows)


def _header_map(ws, row: int = 1) -> dict[str, int]:
    out: dict[str, int] = {}
    for cell in ws[row]:
        if cell.value is None:
            continue
        key = str(cell.value).strip()
        if key:
            out[key] = cell.column
    return out


def _col_letter(hmap: dict[str, int], name: str) -> str:
    c = hmap.get(name)
    if not c:
        raise KeyError(f"ALL_PAIRS missing column {name!r}")
    return get_column_letter(c)


def _inject_all_pairs_leg_pnl_borrow_excel(
    ws,
    *,
    data_start: int = 2,
    data_end: int,
    settings_sheet: str = "dc_workbook_settings",
) -> bool:
    """
    Replace engine ``daily_{long,short,borrow,underlying}`` cells with v15-consistent
    one-row diffs and borrow accrual / ``TRADING_DAYS`` (252 default via settings).

    Requires on ``ALL_PAIRS``: ``etf``, ``under``,
    ``long_sh`` / ``short_sh``, ``underlying_price`` / ``etf_price``,
    ``borrow_rate_annual``, ``under_borrow_rate_annual``,
    ``excel_long_trades_usd`` / ``excel_short_trades_usd`` (cumulative cashflows).

    :returns: whether formulas were written (``False`` if required columns are missing).
    """
    h = _header_map(ws, 1)
    need = (
        "etf",
        "under",
        "long_sh",
        "short_sh",
        "underlying_price",
        "etf_price",
        "borrow_rate_annual",
        "under_borrow_rate_annual",
        "daily_long_pnl_usd",
        "daily_short_pnl_usd",
        "daily_borrow_cost_usd",
        "daily_underlying_borrow_cost_usd",
    )
    for n in need:
        if n not in h:
            return False
    for n in ("excel_long_trades_usd", "excel_short_trades_usd"):
        if n not in h:
            return False
    c_etf = _col_letter(h, "etf")
    c_und = _col_letter(h, "under")
    c_lsh = _col_letter(h, "long_sh")
    c_ssh = _col_letter(h, "short_sh")
    c_pu = _col_letter(h, "underlying_price")
    c_pe = _col_letter(h, "etf_price")
    c_bre = _col_letter(h, "borrow_rate_annual")
    c_bru = _col_letter(h, "under_borrow_rate_annual")
    c_exl = _col_letter(h, "excel_long_trades_usd")
    c_exs = _col_letter(h, "excel_short_trades_usd")
    c_lp = _col_letter(h, "daily_long_pnl_usd")
    c_sp = _col_letter(h, "daily_short_pnl_usd")
    c_bor = _col_letter(h, "daily_borrow_cost_usd")
    c_und_b = _col_letter(h, "daily_underlying_borrow_cost_usd")
    st = re.sub("'", "''", settings_sheet)
    td = f"IFERROR(VLOOKUP(\"trading_days\",'{st}'!$A:$B,2,FALSE),252)"
    for r in range(data_start, data_end + 1):
        r1 = r - 1
        same_pair = f"AND(${c_etf}{r}=${c_etf}{r1},${c_und}{r}=${c_und}{r1})"
        l_now = f"({c_exl}{r}+{c_lsh}{r}*{c_pu}{r})"
        s_now = f"({c_exs}{r}+{c_ssh}{r}*{c_pe}{r})"
        l_prv = f"({c_exl}{r1}+{c_lsh}{r1}*{c_pu}{r1})"
        s_prv = f"({c_exs}{r1}+{c_ssh}{r1}*{c_pe}{r1})"
        ws[f"{c_lp}{r}"] = f"=IF({same_pair},({l_now})-({l_prv}),0)"
        ws[f"{c_sp}{r}"] = f"=IF({same_pair},({s_now})-({s_prv}),0)"
        b_und = f"IF({c_lsh}{r}<0,ABS({c_lsh}{r})*{c_pu}{r}*{c_bru}{r}/({td}),0)"
        b_etf = f"IF({c_ssh}{r}<0,ABS({c_ssh}{r})*{c_pe}{r}*{c_bre}{r}/({td}),0)"
        ws[f"{c_und_b}{r}"] = f"={b_und}"
        ws[f"{c_bor}{r}"] = f"=({b_etf})+({b_und})"
    return True


def _inject_all_pairs_formulas(ws, *, data_start: int = 2, data_end: int) -> None:
    h = _header_map(ws, 1)
    c_lsh = _col_letter(h, "long_sh")
    c_ssh = _col_letter(h, "short_sh")
    c_pu = _col_letter(h, "underlying_price")
    c_pe = _col_letter(h, "etf_price")
    c_lmb = _col_letter(h, "long_margin_basis_usd")
    c_sfb = _col_letter(h, "short_financing_basis_usd")
    c_ln = _col_letter(h, "long_notional_usd")
    c_sn = _col_letter(h, "short_notional_usd")
    c_gr = _col_letter(h, "gross_notional_usd")
    c_nt = _col_letter(h, "net_notional_usd")
    c_lp = _col_letter(h, "daily_long_pnl_usd")
    c_sp = _col_letter(h, "daily_short_pnl_usd")
    c_bor = _col_letter(h, "daily_borrow_cost_usd")
    c_ubor = _col_letter(h, "daily_underlying_borrow_cost_usd")
    c_mar = get_column_letter(h["daily_margin_debit_cost_usd"]) if "daily_margin_debit_cost_usd" in h else None
    c_sci = get_column_letter(h["daily_short_credit_income_usd"]) if "daily_short_credit_income_usd" in h else None
    c_nfin = get_column_letter(h["daily_net_financing_cost_usd"]) if "daily_net_financing_cost_usd" in h else None
    c_txn = get_column_letter(h["daily_txn_cost_usd"]) if "daily_txn_cost_usd" in h else None
    c_grossp = _col_letter(h, "daily_pair_gross_trading_pnl_usd")
    c_netp = _col_letter(h, "daily_pair_net_pnl_usd")
    c_extxn = _col_letter(h, "daily_pair_net_ex_txn_usd")

    for r in range(data_start, data_end + 1):
        d, e, f, g = c_lsh, c_ssh, c_pu, c_pe
        ws[f"{c_lmb}{r}"] = f"=MAX({d}{r},0)*{f}{r}"
        ws[f"{c_sfb}{r}"] = f"=MAX(-{d}{r},0)*{f}{r}+MAX(-{e}{r},0)*{g}{r}"
        ws[f"{c_ln}{r}"] = f"={d}{r}*{f}{r}"
        ws[f"{c_sn}{r}"] = f"={e}{r}*{g}{r}"
        ws[f"{c_gr}{r}"] = f"=ABS({d}{r})*{f}{r}+ABS({e}{r})*{g}{r}"
        ws[f"{c_nt}{r}"] = f"={c_ln}{r}+{c_sn}{r}"
        if c_mar:
            ws[f"{c_mar}{r}"] = 0.0
        if c_sci:
            ws[f"{c_sci}{r}"] = 0.0
        if c_nfin:
            ws[f"{c_nfin}{r}"] = 0.0
        if c_txn:
            ws[f"{c_txn}{r}"] = 0.0
        ws[f"{c_grossp}{r}"] = f"={c_lp}{r}+{c_sp}{r}"
        ws[f"{c_netp}{r}"] = f"={c_grossp}{r}-{c_bor}{r}-{c_ubor}{r}"
        ws[f"{c_extxn}{r}"] = f"={c_netp}{r}"


def _inject_all_pairs_pnl_net_of_borrow(ws, *, data_start: int, data_end: int) -> None:
    """``pnl_net_of_borrow_usd`` = leg P&L sum minus cumulative borrow (values in ``cum_borrow_cost_usd``)."""
    h = _header_map(ws, 1)
    if "pnl_net_of_borrow_usd" not in h or "cum_borrow_cost_usd" not in h:
        return
    c_lp = get_column_letter(h["daily_long_pnl_usd"])
    c_sp = get_column_letter(h["daily_short_pnl_usd"])
    c_cb = get_column_letter(h["cum_borrow_cost_usd"])
    c_out = get_column_letter(h["pnl_net_of_borrow_usd"])
    for r in range(data_start, data_end + 1):
        ws[f"{c_out}{r}"] = f"={c_lp}{r}+{c_sp}{r}-{c_cb}{r}"


def _inject_portfolio_ledger_dc_formulas(
    wb: Workbook,
    *,
    n_data_rows: int,
    h_ap: dict[str, int],
    all_pairs_sheet: str = "ALL_PAIRS",
    settings_sheet: str = "dc_workbook_settings",
    book_sheet: str = "book_daily",
) -> None:
    """Reference-style nine-column ``portfolio_ledger_dc`` (variable pair counts via ``SUMIFS``)."""
    if n_data_rows <= 0 or "portfolio_ledger_dc" not in wb.sheetnames:
        return
    if settings_sheet not in wb.sheetnames:
        return
    ws = wb["portfolio_ledger_dc"]
    ap = re.sub("'", "''", all_pairs_sheet)
    st = re.sub("'", "''", settings_sheet)
    bd = re.sub("'", "''", book_sheet)

    def _L(name: str) -> str:
        cix = h_ap.get(name)
        if not cix:
            raise KeyError(f"ALL_PAIRS sheet missing column {name!r} (needed for portfolio_ledger_dc)")
        return get_column_letter(cix)

    c_dt = _L("date")
    c_ln = _L("long_notional_usd")
    c_pnb = _L("pnl_net_of_borrow_usd")

    for r in range(2, n_data_rows + 2):
        a = f"A{r}"
        crit = f"'{ap}'!${c_dt}:${c_dt},{a}"
        ws[f"B{r}"] = f"=SUMIFS('{ap}'!${c_ln}:${c_ln},{crit})"
        ws[f"C{r}"] = f"='{st}'!$B$2"
        ws[f"D{r}"] = f"=IF(ROW()=2,0,$B{r}*$C{r}/'{st}'!$B$1)"
        ws[f"E{r}"] = f"=IFERROR(VLOOKUP({a},'{bd}'!$A:$C,3,FALSE),0)"
        ws[f"F{r}"] = f"=D{r}+E{r}"
        ws[f"G{r}"] = f"=SUMIFS('{ap}'!${c_pnb}:${c_pnb},{crit})"
        ws[f"H{r}"] = f"=G{r}-F{r}"
        if r == 2:
            ws[f"I{r}"] = f"=H{r}"
        else:
            ws[f"I{r}"] = f"=H{r}-H{r-1}"


def _inject_portfolio_ledger_formulas(
    ws_pl,
    *,
    n_data_rows: int,
    h_ap: dict[str, int],
    all_pairs_sheet: str = "ALL_PAIRS",
    book_sheet: str = "book_daily",
    financing_daycount: float = 360.0,
) -> None:
    """
    Portfolio roll-up: ``SUMIFS`` / ``AVERAGEIFS`` into ``ALL_PAIRS`` using the
    workbook's row-1 headers (so column letters track the export schema).

    ``daily_portfolio_margin_usd`` uses ``financing_daycount`` (default 360,
    matching ``CFG['financing_daycount']`` in the Diamond Creek notebooks), not
    ``252``.

    ``net_after…`` is ``N - F - G`` where ``G`` is portfolio-level daily
    transaction cost pulled from ``book_daily``.
    """
    if n_data_rows <= 0:
        return
    ap = re.sub("'", "''", all_pairs_sheet)
    bd = re.sub("'", "''", book_sheet)

    def _L(name: str) -> str:
        cix = h_ap.get(name)
        if not cix:
            raise KeyError(f"ALL_PAIRS sheet missing column {name!r} (needed for portfolio formulas)")
        return get_column_letter(cix)

    c_ad = _L("date")
    c_lmb = _L("long_margin_basis_usd")
    c_ff = _L("fed_funds_rate")
    c_lp = _L("daily_long_pnl_usd")
    c_sp = _L("daily_short_pnl_usd")
    c_net = _L("daily_pair_net_pnl_usd")
    dc = float(financing_daycount)
    dc_s = str(int(dc)) if dc == int(dc) else str(dc)

    for r in range(2, n_data_rows + 2):
        a = f"A{r}"
        crit = f"'{ap}'!${c_ad}:${c_ad},{a}"
        # A date, D margin_spread (values) — do not overwrite D.
        ws_pl[f"B{r}"] = f"=SUMIFS('{ap}'!${c_lmb}:${c_lmb},{crit})"
        ws_pl[f"C{r}"] = f"=AVERAGEIFS('{ap}'!${c_ff}:${c_ff},{crit})"
        ws_pl[f"E{r}"] = f"=C{r}+D{r}"
        ws_pl[f"F{r}"] = f"=B{r}*E{r}/{dc_s}"
        ws_pl[f"G{r}"] = f"=IFERROR(VLOOKUP({a},'{bd}'!$A:$C,3,FALSE),0)"
        ws_pl[f"H{r}"] = f"=SUMIFS('{ap}'!${c_lp}:${c_lp},{crit})"
        ws_pl[f"I{r}"] = f"=SUMIFS('{ap}'!${c_sp}:${c_sp},{crit})"
        ws_pl[f"J{r}"] = f"=H{r}+I{r}"
        ws_pl[f"K{r}"] = f"=IFERROR(VLOOKUP({a},'{bd}'!$A:$G,3,FALSE),0)"
        ws_pl[f"L{r}"] = f"=IFERROR(VLOOKUP({a},'{bd}'!$A:$G,7,FALSE),0)"
        ws_pl[f"M{r}"] = f"=K{r}+L{r}"
        ws_pl[f"N{r}"] = f"=SUMIFS('{ap}'!${c_net}:${c_net},{crit})"
        ws_pl[f"O{r}"] = f"=N{r}-F{r}-G{r}"
        if r == 2:
            ws_pl[f"P{r}"] = f"=J{r}"
            ws_pl[f"Q{r}"] = f"=N{r}"
        else:
            ws_pl[f"P{r}"] = f"=P{r-1}+J{r}"
            ws_pl[f"Q{r}"] = f"=Q{r-1}+N{r}"
        ws_pl[f"R{r}"] = f"=IFERROR(VLOOKUP({a},'{bd}'!$A:$G,4,FALSE),0)"
        ws_pl[f"S{r}"] = f"=R{r}"
        ws_pl[f"U{r}"] = f"=IF(ISNUMBER(T{r}),O{r}-T{r},\"\")"


def _inject_checks_sheet(
    wb: Workbook,
    *,
    n_check_rows: int,
    h_ap: dict[str, int],
    all_pairs_sheet: str = "ALL_PAIRS",
    book_sheet: str = "book_daily",
) -> None:
    """Build ``checks`` sheet: date-level pair sums vs ``book_daily``."""
    if n_check_rows <= 0:
        return
    if "checks" not in wb.sheetnames:
        wb.create_sheet("checks")
    ws = wb["checks"]
    ap = re.sub("'", "''", all_pairs_sheet)
    bd = re.sub("'", "''", book_sheet)

    def _ac(name: str) -> str:
        cix = h_ap.get(name)
        if not cix:
            raise KeyError(f"ALL_PAIRS missing {name!r} (needed for checks formulas)")
        return get_column_letter(cix)

    c_ad = _ac("date")
    c_lp = _ac("daily_long_pnl_usd")
    c_sp = _ac("daily_short_pnl_usd")
    c_bor = _ac("daily_borrow_cost_usd")
    c_ext = _ac("daily_pair_net_ex_txn_usd")
    # Layout row1: labels row2+ data
    headers = [
        "date",
        "sum_pair_long_pnl_usd",
        "sum_pair_short_pnl_usd",
        "sum_pair_gross_ls_usd",
        "book_daily_long_short_pnl_usd",
        "diff_gross_usd",
        "sum_pair_borrow_usd",
        "book_daily_borrow_usd",
        "diff_borrow_usd",
        "sum_pair_txn_usd",
        "book_daily_txn_usd",
        "diff_txn_usd",
        "sum_pair_ex_txn_net_usd",
        "book_nav_change_plus_txn_usd",
        "diff_attr_vs_book_net_usd",
    ]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    for r in range(2, n_check_rows + 2):
        ar = f"A{r}"
        cr = f"'{ap}'!${c_ad}:${c_ad},{ar}"
        ws[f"B{r}"] = f"=SUMIFS('{ap}'!${c_lp}:${c_lp},{cr})"
        ws[f"C{r}"] = f"=SUMIFS('{ap}'!${c_sp}:${c_sp},{cr})"
        ws[f"D{r}"] = f"=B{r}+C{r}"
        ws[f"E{r}"] = f"=IFERROR(VLOOKUP({ar},'{bd}'!$A:$F,6,FALSE),0)"
        ws[f"F{r}"] = f"=D{r}-E{r}"
        ws[f"G{r}"] = f"=SUMIFS('{ap}'!${c_bor}:${c_bor},{cr})"
        ws[f"H{r}"] = f"=IFERROR(VLOOKUP({ar},'{bd}'!$A:$F,5,FALSE),0)"
        ws[f"I{r}"] = f"=G{r}-H{r}"
        ws[f"J{r}"] = f"=IFERROR(VLOOKUP({ar},'{bd}'!$A:$C,3,FALSE),0)"
        ws[f"K{r}"] = f"=IFERROR(VLOOKUP({ar},'{bd}'!$A:$C,3,FALSE),0)"
        ws[f"L{r}"] = f"=J{r}-K{r}"
        ws[f"M{r}"] = f"=SUMIFS('{ap}'!${c_ext}:${c_ext},{cr})"
        ws[f"N{r}"] = f"=IFERROR(VLOOKUP({ar},'{bd}'!$A:$F,4,FALSE),0)+IFERROR(VLOOKUP({ar},'{bd}'!$A:$C,3,FALSE),0)"
        ws[f"O{r}"] = f"=M{r}-N{r}"


def _inject_per_pair_sumifs(wb: Workbook, *, skip: Iterable[str]) -> None:
    """For each pair sheet, rewrite key numeric columns to ``SUMIFS`` against ``ALL_PAIRS``."""
    ap = "ALL_PAIRS"
    ap_q = re.sub("'", "''", ap)
    h_ap = _header_map(wb["ALL_PAIRS"], 1)
    c_ap_date = get_column_letter(h_ap["date"])
    c_ap_etf = get_column_letter(h_ap["etf"])
    pair_sumif_fields = (
        "long_notional_usd",
        "short_notional_usd",
        "long_margin_basis_usd",
        "short_financing_basis_usd",
        "gross_notional_usd",
        "net_notional_usd",
        "daily_long_pnl_usd",
        "daily_short_pnl_usd",
        "daily_borrow_cost_usd",
        "daily_underlying_borrow_cost_usd",
        "daily_margin_debit_cost_usd",
        "daily_short_credit_income_usd",
        "daily_net_financing_cost_usd",
        "daily_txn_cost_usd",
        "daily_turnover_usd",
        "daily_pair_gross_trading_pnl_usd",
        "daily_pair_net_pnl_usd",
        "daily_pair_net_ex_txn_usd",
    )
    col_allpairs = {f: get_column_letter(h_ap[f]) for f in pair_sumif_fields if f in h_ap}
    skip_set = set(skip) | {"ALL_PAIRS", "portfolio_ledger", "fund_vs_sim_daily", "book_daily", "checks"}
    for name in wb.sheetnames:
        if name in skip_set:
            continue
        ws = wb[name]
        h = _header_map(ws, 1)
        if "date" not in h or "etf" not in h:
            continue
        c_a = get_column_letter(h["date"])
        c_b = get_column_letter(h["etf"])
        max_r = ws.max_row or 1
        for r in range(2, max_r + 1):
            for field, ap_col in col_allpairs.items():
                if field not in h:
                    continue
                c = get_column_letter(h[field])
                ws[f"{c}{r}"] = (
                    f"=SUMIFS('{ap_q}'!${ap_col}:${ap_col},'{ap_q}'!${c_ap_date}:${c_ap_date},{c_a}{r},"
                    f"'{ap_q}'!${c_ap_etf}:${c_ap_etf},{c_b}{r})"
                )
            if "cum_borrow_cost_usd" in h and "daily_borrow_cost_usd" in h:
                c_cum = get_column_letter(h["cum_borrow_cost_usd"])
                c_bor = get_column_letter(h["daily_borrow_cost_usd"])
                # ``Diamond_Creek_Daily_Pairs`` per-pair style: ``=SUM($L$2:L2)``.
                ws[f"{c_cum}{r}"] = f"=SUM(${c_bor}$2:{c_bor}{r})"


def inject_daily_pair_workbook_formulas(
    path: Path | str,
    *,
    all_pairs_data_rows: int,
    portfolio_data_rows: int,
    checks_data_rows: int,
    inject_per_pair_sumifs: bool = True,
    inject_leg_pnl_borrow: bool = True,
    verbose: bool = True,
    financing_daycount: float = 360.0,
    include_portfolio_ledger_dc: bool = True,
) -> None:
    """
    Post-process workbook written by ``export_diamond_creek_daily_pair_workbook``.

    :param all_pairs_data_rows: number of data rows (excluding header) on ``ALL_PAIRS``.
    :param portfolio_data_rows: number of data rows on ``portfolio_ledger``.
    :param checks_data_rows: should match portfolio (one row per date); if 0, skip checks body.
    :param inject_leg_pnl_borrow: when ``True`` and trade-cash columns + rates exist, rewrite
        per-row leg P&L and borrow.  Set ``False`` when the caller has pre-filled borrow from
        a book-level or net-by-under model (e.g. ETF arb pairwise export).
    :param financing_daycount: accrual divisor for the simplified portfolio margin column
        ``daily_portfolio_margin_usd`` (``B*E/daycount``); default ``360`` matches the
        Diamond Creek notebook ``CFG['financing_daycount']``.
    :param include_portfolio_ledger_dc: when ``True`` (default), write ``pnl_net_of_borrow_usd``
        on ``ALL_PAIRS`` and formulas on ``portfolio_ledger_dc`` / ``dc_workbook_settings``.
    """
    p = Path(path)
    wb = load_workbook(p)
    if "ALL_PAIRS" not in wb.sheetnames:
        if verbose:
            print(f"[daily_pair_workbook_formulas] skip: no ALL_PAIRS sheet in {p}")
        wb.close()
        return
    ws_ap = wb["ALL_PAIRS"]
    h_ap = _header_map(ws_ap, 1)
    if all_pairs_data_rows > 0:
        if inject_leg_pnl_borrow and _inject_all_pairs_leg_pnl_borrow_excel(
            ws_ap, data_start=2, data_end=1 + all_pairs_data_rows
        ):
            h_ap = _header_map(ws_ap, 1)
        _inject_all_pairs_formulas(ws_ap, data_start=2, data_end=1 + all_pairs_data_rows)
        h_ap = _header_map(ws_ap, 1)
        if include_portfolio_ledger_dc:
            _inject_all_pairs_pnl_net_of_borrow(ws_ap, data_start=2, data_end=1 + all_pairs_data_rows)
            h_ap = _header_map(ws_ap, 1)
    if "portfolio_ledger" in wb.sheetnames and portfolio_data_rows > 0:
        if "book_daily" not in wb.sheetnames:
            if verbose:
                print(
                    f"[daily_pair_workbook_formulas] warning: no 'book_daily' sheet — "
                    f"skipping portfolio_ledger / checks formulas (see build_book_daily_dataframe)."
                )
        else:
            _inject_portfolio_ledger_formulas(
                wb["portfolio_ledger"],
                n_data_rows=portfolio_data_rows,
                h_ap=h_ap,
                financing_daycount=financing_daycount,
            )
    if (
        include_portfolio_ledger_dc
        and portfolio_data_rows > 0
        and "portfolio_ledger_dc" in wb.sheetnames
        and all_pairs_data_rows > 0
    ):
        _inject_portfolio_ledger_dc_formulas(
            wb,
            n_data_rows=portfolio_data_rows,
            h_ap=h_ap,
        )
    if checks_data_rows > 0 and "book_daily" in wb.sheetnames:
        _inject_checks_sheet(wb, n_check_rows=checks_data_rows, h_ap=h_ap)
    if inject_per_pair_sumifs:
        _inject_per_pair_sumifs(wb, skip=SKIP_SHEETS)
    wb.save(p)
    wb.close()
    if verbose:
        extra = ", portfolio_ledger_dc" if include_portfolio_ledger_dc else ""
        print(
            f"[daily_pair_workbook_formulas] wrote formulas: ALL_PAIRS rows={all_pairs_data_rows}, "
            f"portfolio rows={portfolio_data_rows}, checks rows={checks_data_rows}{extra} -> {p}"
        )
