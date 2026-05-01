"""Compare NAV/exposure behavior between NEW pairwise workbook and REF."""
from openpyxl import load_workbook
import datetime as dt

NEW_P = r"C:\Users\werdn\Downloads\DC ETF Arb Pairwise Backtest Attribution.xlsx"
REF_P = r"C:\Users\werdn\Downloads\DC ETF Arb Attribution.xlsx"

print("=== NEW: book_daily ===")
wb = load_workbook(NEW_P, data_only=True)
ws = wb["book_daily"]
print("shape:", ws.max_row, "x", ws.max_column)
for r in range(1, min(6, ws.max_row + 1)):
    print(r, [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])
print("...")
for r in (ws.max_row - 3, ws.max_row - 2, ws.max_row - 1, ws.max_row):
    print(r, [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

# Sample first/last NAV
ws_d = wb["Daily PnL"]
print()
print("=== NEW: Daily PnL header ===")
for r in range(1, 6):
    print(r, [ws_d.cell(row=r, column=c).value for c in range(1, min(ws_d.max_column, 8) + 1)])

# Pull NAV-equivalent: sum monthly Net PnL through time
print()
print("=== NEW: Monthly Net cumsum (book NAV) ===")
ws_m = wb["Monthly Attribution"]
cum = 10_000_000.0  # base capital
for r in range(4, ws_m.max_row + 1):
    m = ws_m[f"B{r}"].value
    n = ws_m[f"T{r}"].value
    if m is None: continue
    if isinstance(n, (int, float)):
        cum += n
    f_long = ws_m[f"F{r}"].value
    g_short = ws_m[f"G{r}"].value
    if isinstance(f_long, (int, float)) and isinstance(cum, float):
        long_to_nav = f_long / cum if cum else 0
        print(f"{m}  NAV~{cum:>14,.0f}  Long={f_long:>14,.0f} ({long_to_nav:>5.2f}x)  Short={g_short:>14,.0f}")

print()
print("=== REF: Monthly Net cumsum (book NAV) ===")
wb2 = load_workbook(REF_P, data_only=True)
ws_mr = wb2["Monthly Attribution"]
cum2 = 10_000_000.0
for r in range(4, ws_mr.max_row + 1):
    m = ws_mr[f"B{r}"].value
    n = ws_mr[f"T{r}"].value
    if m is None: continue
    if isinstance(n, (int, float)):
        cum2 += n
    f_long = ws_mr[f"F{r}"].value
    g_short = ws_mr[f"G{r}"].value
    if isinstance(f_long, (int, float)):
        ratio = f_long / cum2 if cum2 else 0
        print(f"{m}  NAV~{cum2:>14,.0f}  Long={f_long:>14,.0f} ({ratio:>5.2f}x)  Short={g_short:>14,.0f}")
