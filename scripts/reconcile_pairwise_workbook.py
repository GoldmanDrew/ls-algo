"""
Light-weight checks on ``DC ETF Arb Pairwise Backtest Attribution.xlsx``.

Usage:
  python scripts/reconcile_pairwise_workbook.py path/to/workbook.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("workbook", type=Path, help="Pairwise attribution .xlsx")
    args = p.parse_args()
    wb = load_workbook(args.workbook, data_only=True)
    try:
        if "LP_FEE" not in wb.sheetnames:
            print("ERROR: missing LP_FEE sheet", file=sys.stderr)
            return 2
        ws = wb["LP_FEE"]
        perf_sum = 0.0
        last_date = None
        last_perf = None
        for r in range(2, ws.max_row + 1):
            d = ws.cell(r, 1).value
            perf = ws.cell(r, 3).value
            if isinstance(perf, (int, float)) and perf != 0:
                perf_sum += float(perf)
            if d is not None:
                last_date = d
                last_perf = float(perf or 0) if isinstance(perf, (int, float)) else perf
        print("LP_FEE: sum(perf_usd) =", f"{perf_sum:,.2f}")
        print("LP_FEE: last row date =", last_date, " perf_usd =", last_perf)
        if last_perf is not None and float(last_perf) <= 0 and last_date is not None:
            y = getattr(last_date, "year", None)
            m = getattr(last_date, "month", None)
            if y == 2026 and m == 3:
                print(
                    "WARN: last row is 2026-03 but perf_usd is 0 — "
                    "re-export with crystallize_trailing_partial_year=True (default in export).",
                    file=sys.stderr,
                )
    finally:
        wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
