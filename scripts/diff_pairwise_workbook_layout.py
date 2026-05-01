"""
Compare column widths and freeze panes between two workbooks (same sheet names).

Usage:
  python scripts/diff_pairwise_workbook_layout.py golden.xlsx candidate.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


def _dims(ws):
    out = {}
    for k, v in ws.column_dimensions.items():
        if v is None or v.width is None:
            continue
        out[k] = (v.width, bool(getattr(v, "hidden", False)))
    return out


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("golden", type=Path)
    p.add_argument("candidate", type=Path)
    args = p.parse_args()
    g = load_workbook(args.golden, data_only=False)
    c = load_workbook(args.candidate, data_only=False)
    rc = 0
    try:
        for name in g.sheetnames:
            if name not in c.sheetnames:
                print(f"MISSING sheet in candidate: {name}")
                rc = 1
                continue
            gs, cs = g[name], c[name]
            if gs.freeze_panes != cs.freeze_panes:
                print(f"{name}: freeze_panes golden={gs.freeze_panes!r} candidate={cs.freeze_panes!r}")
                rc = 1
            dg, dc = _dims(gs), _dims(cs)
            for col in sorted(set(dg) | set(dc), key=lambda x: (len(x), x)):
                if dg.get(col) != dc.get(col):
                    print(f"{name} col {col}: golden={dg.get(col)} candidate={dc.get(col)}")
                    rc = 1
        if rc == 0:
            print("OK: layout dimensions match for common sheets.")
    finally:
        g.close()
        c.close()
    return rc


if __name__ == "__main__":
    raise SystemExit(main())
