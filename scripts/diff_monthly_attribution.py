"""Diff Monthly Attribution between reference and new pairwise workbooks."""
from openpyxl import load_workbook

REF_P = r"C:\Users\werdn\Downloads\DC ETF Arb Attribution.xlsx"
NEW_P = r"C:\Users\werdn\Downloads\DC ETF Arb Pairwise Backtest Attribution.xlsx"

COLS = ["B", "D", "F", "G", "H", "I", "K", "L", "M", "N", "P", "Q", "R", "T"]
NAMES = [
    "Month", "Bench", "LongN", "ShortN", "PnL_L", "PnL_S",
    "TCost", "Borrow", "Margin", "PreFee", "Mgmt", "Inc", "Tot", "Net",
]


def load(p):
    wb = load_workbook(p, data_only=True)
    ws = wb["Monthly Attribution"]
    rows = []
    for r in range(4, ws.max_row + 1):
        rec = {}
        for c, n in zip(COLS, NAMES):
            rec[n] = ws[f"{c}{r}"].value
        if rec["Month"] is None:
            continue
        rows.append(rec)
    return rows


def f(x):
    if x is None:
        return ""
    if isinstance(x, (int, float)):
        return f"{x:,.0f}"
    return str(x)


def pct(a, b):
    if a in (None, 0) or b is None:
        return ""
    return f"{(b-a)/abs(a)*100:+.1f}%"


def df(a, b):
    if a is None or b is None:
        return ""
    return f"{(b-a):,.0f}"


def main():
    ref = load(REF_P)
    new = load(NEW_P)
    ref_map = {r["Month"]: r for r in ref}
    new_map = {r["Month"]: r for r in new}
    months = sorted(set(list(ref_map.keys()) + list(new_map.keys())))

    print(
        f"{'Month':>8} | "
        f"{'LongN_REF':>12} {'LongN_NEW':>12} {'dLong%':>7} | "
        f"{'ShortN_REF':>12} {'ShortN_NEW':>12} {'dSh%':>7} | "
        f"{'PnL_L_REF':>11} {'PnL_L_NEW':>11} {'PnL_S_REF':>11} {'PnL_S_NEW':>11} | "
        f"{'PreFee_REF':>12} {'PreFee_NEW':>12} {'dPreFee':>10} | "
        f"{'Mgmt_REF':>9} {'Mgmt_NEW':>9} {'Inc_REF':>9} {'Inc_NEW':>9} | "
        f"{'Net_REF':>12} {'Net_NEW':>12}"
    )
    for m in months:
        r = ref_map.get(m, {n: None for n in NAMES})
        n = new_map.get(m, {n: None for n in NAMES})
        print(
            f"{str(m):>8} | "
            f"{f(r['LongN']):>12} {f(n['LongN']):>12} {pct(r['LongN'], n['LongN']):>7} | "
            f"{f(r['ShortN']):>12} {f(n['ShortN']):>12} {pct(r['ShortN'], n['ShortN']):>7} | "
            f"{f(r['PnL_L']):>11} {f(n['PnL_L']):>11} {f(r['PnL_S']):>11} {f(n['PnL_S']):>11} | "
            f"{f(r['PreFee']):>12} {f(n['PreFee']):>12} {df(r['PreFee'], n['PreFee']):>10} | "
            f"{f(r['Mgmt']):>9} {f(n['Mgmt']):>9} {f(r['Inc']):>9} {f(n['Inc']):>9} | "
            f"{f(r['Net']):>12} {f(n['Net']):>12}"
        )

    print()
    print("ROW-LEVEL TOTALS:")
    for k in ["LongN", "ShortN", "PnL_L", "PnL_S", "TCost", "Borrow", "Margin", "PreFee", "Mgmt", "Inc", "Net"]:
        rs = sum((r.get(k) or 0) for r in ref)
        ns = sum((r.get(k) or 0) for r in new)
        print(f"  {k:>8}: REF={rs:>16,.0f} NEW={ns:>16,.0f} diff={ns-rs:>14,.0f} ({(ns-rs)/abs(rs)*100 if rs else 0:+.2f}%)")


if __name__ == "__main__":
    main()
