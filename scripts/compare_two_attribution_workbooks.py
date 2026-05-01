"""High-level compare Monthly Attribution + Daily PnL between two DC workbooks."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


PAIR = Path(r"c:\Users\werdn\Downloads\DC ETF Arb Pairwise Backtest Attribution.xlsx")
REF = Path(r"c:\Users\werdn\Downloads\DC ETF Arb Attribution.xlsx")


def _monthly(ws) -> list[dict]:
    rows = []
    for r in range(4, ws.max_row + 1):
        m = ws[f"B{r}"].value
        if m is None or str(m).strip() == "":
            continue
        rows.append(
            {
                "month": str(m),
                "bench": ws[f"D{r}"].value,
                "long_n": ws[f"F{r}"].value,
                "short_n": ws[f"G{r}"].value,
                "gross_l": ws[f"H{r}"].value,
                "gross_s": ws[f"I{r}"].value,
                "tcost": ws[f"K{r}"].value,
                "borrow": ws[f"L{r}"].value,
                "margin": ws[f"M{r}"].value,
                "pre_fee": ws[f"N{r}"].value,
                "mgmt": ws[f"P{r}"].value,
                "inc": ws[f"Q{r}"].value,
                "fee_tot": ws[f"R{r}"].value,
                "net": ws[f"T{r}"].value,
            }
        )
    return rows


def _sum_col(rows, key):
    return sum(float(x[key] or 0) for x in rows if isinstance(x.get(key), (int, float)))


def _daily_sum(ws, col_letter: str = "C") -> float:
    """Daily PnL sheet: assume header row 2, data from row 3."""
    s = 0.0
    for r in range(3, ws.max_row + 1):
        v = ws[f"{col_letter}{r}"].value
        if isinstance(v, (int, float)):
            s += float(v)
    return s


def _daily_rows(ws) -> int:
    n = 0
    for r in range(3, ws.max_row + 1):
        if ws[f"B{r}"].value is not None:
            n += 1
    return n


def main() -> None:
    for p in (PAIR, REF):
        if not p.is_file():
            raise SystemExit(f"Missing file: {p}")

    wb_p = load_workbook(PAIR, data_only=True)
    wb_r = load_workbook(REF, data_only=True)

    mp = _monthly(wb_p["Monthly Attribution"])
    mr = _monthly(wb_r["Monthly Attribution"])

    print("=== FILES ===")
    print("PAIR:", PAIR)
    print("REF :", REF)
    print()

    print("=== SHEETS ===")
    print("PAIR sheets:", wb_p.sheetnames)
    print("REF  sheets:", wb_r.sheetnames)
    print()

    print("=== MONTHLY ROWS ===")
    print(f"PAIR months: {len(mp)}  REF months: {len(mr)}")
    print()

    keys = [
        "long_n",
        "short_n",
        "gross_l",
        "gross_s",
        "tcost",
        "borrow",
        "margin",
        "pre_fee",
        "mgmt",
        "inc",
        "net",
    ]
    labels = {
        "long_n": "Long notional (sum F)",
        "short_n": "Short notional (sum G, signed)",
        "gross_l": "Gross PnL long (H)",
        "gross_s": "Gross PnL short (I)",
        "tcost": "T-costs (K)",
        "borrow": "Borrow (L)",
        "margin": "Margin (M)",
        "pre_fee": "Pre-fee PnL (N)",
        "mgmt": "Mgmt (P)",
        "inc": "Incentive (Q)",
        "net": "Net PnL (T)",
    }

    print("=== COLUMN TOTALS (Monthly sum) ===")
    print(f"{'metric':<28} {'REF':>16} {'PAIR':>16} {'PAIR-REF':>14}")
    for k in keys:
        a, b = _sum_col(mr, k), _sum_col(mp, k)
        print(f"{labels[k]:<28} {a:>16,.0f} {b:>16,.0f} {b - a:>14,.0f}")
    print()

    # Gross / implied NAV from cumulative net (rough)
    cum_r = cum_p = 10_000_000.0
    gnav_r = gnav_p = []
    for r, p in zip(mr, mp):
        if r["month"] != p["month"]:
            print("WARN: month misalignment", r["month"], p["month"])
            break
        if isinstance(r["net"], (int, float)):
            cum_r += float(r["net"])
        if isinstance(p["net"], (int, float)):
            cum_p += float(p["net"])
        ln_r, ln_p = r["long_n"], p["long_n"]
        if isinstance(ln_r, (int, float)) and cum_r > 0:
            gnav_r.append((abs(float(r["short_n"] or 0)) + float(ln_r)) / cum_r)
        if isinstance(ln_p, (int, float)) and cum_p > 0:
            gnav_p.append((abs(float(p["short_n"] or 0)) + float(ln_p)) / cum_p)

    def _avg(xs):
        return sum(xs) / len(xs) if xs else float("nan")

    print("=== IMPLIED GROSS / NAV (month-end long+|short| over cum net+10M) ===")
    print(f"REF  mean gross/NAV: {_avg(gnav_r):.3f}x  (min {min(gnav_r):.3f}, max {max(gnav_r):.3f})")
    print(f"PAIR mean gross/NAV: {_avg(gnav_p):.3f}x  (min {min(gnav_p):.3f}, max {max(gnav_p):.3f})")
    print()

    # Daily PnL
    d_p = wb_p["Daily PnL"]
    d_r = wb_r["Daily PnL"]
    print("=== DAILY PnL ===")
    print(f"REF  rows: {_daily_rows(d_r)}  sum col C: {_daily_sum(d_r, 'C'):,.0f}")
    print(f"PAIR rows: {_daily_rows(d_p)}  sum col C: {_daily_sum(d_p, 'C'):,.0f}")
    print()

    wb_p.close()
    wb_r.close()


if __name__ == "__main__":
    main()
