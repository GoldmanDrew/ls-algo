"""
================================================================================
IMPLEMENTATION PROMPT (spec) — ``DC ETF Arb Pairwise Backtest Attribution.xlsx``
================================================================================
**Goal:** Emit a workbook that is **layout-identical** to
``DC ETF Arb Attribution.xlsx`` (same sheet names, merged header cells, column
order, column widths, and Excel formulas on **Monthly Attribution**), but whose
**numeric body** is driven by the **pair backtest + book** instead of the fund’s
published history.

With ``use_excel_formulas=True`` (default), the workbook also embeds ``ALL_PAIRS``,
``book_raw`` (values), ``book_daily`` (formulas to ``book_raw``), and ``LP_FEE``.
**Management** in ``LP_FEE!B`` is **Excel**
(``=IF(include_fees_off, 0, IF(A=F, (mgmt_q)*E, 0))``); the helpers
**D=q_first_trd**, **E=prev-quarter close NAV** (mgmt base), **F=q_last_trd**
are **Python-precomputed values** so the formula does not depend on
``MAXIFS``/``MINIFS`` availability and uses exact date equality. So it lands on
the last trading day of Mar/Jun/Sep/Dec.
**Performance** in ``LP_FEE!C`` is still **values** from ``lp_fees_v15`` (pass-2 annual allocation).
Per-pair **margin** / **short-credit** / **net financing** / **txn** in ``ALL_PAIRS`` keep
**engine** values (not the portfolio-zeros path), and daily pair net/Ex-txn formulas
subtract borrow, net financing, and txn accordingly.
Leg PnL / borrow
/ notionals on ``ALL_PAIRS`` are formula-driven
(``inject_daily_pair_workbook_formulas``). ``DAILY_CALC`` looks up
``mgmt_usd``/``perf_usd`` from ``LP_FEE``.

**Notional and “gross” books (single definition used here)**

* **Signed leg notionals (``ALL_PAIRS``):** ``long_notional_usd = long_sh *``
  ``underlying_price``; ``short_notional_usd = short_sh * etf_price``. Use these
  for *signed* economics and for ``gross_notional_usd = ABS(long_sh)*PU+ABS(ssh)*ETF``
  after ``_inject_all_pairs_formulas`` (see ``scripts/daily_pair_workbook_formulas.py``).
* **Monthly Attribution column map (matches the template — verified against the
  template workbook’s headers/merges):**

  =====  =============================================  ============================================
  Col    Header                                         Body
  =====  =============================================  ============================================
  B      Month                                          ``yyyy-mm`` string
  D      Benchmark Rate                                 ``AVERAGEIFS(ALL_PAIRS!fed_funds...)``
  F      Notional / Long                                ``SUMIFS(long_notional_usd, A=$V{r})``
  G      Notional / Short                               ``SUMIFS(short_notional_usd, A=$V{r})``
  H      Gross PnL / Long                               ``SUMIFS(daily_long_pnl_usd, month range)``
  I      Gross PnL / Short                              ``SUMIFS(daily_short_pnl_usd, month range)``
  K      T-Costs                                        ``SUMIFS(book_daily!C, month range)``
  L      Short Book Borrow Cost                         ``SUMIFS(daily_borrow_usd, month range)``
  M      Long Book Margin Cost                          ``SUMIFS(book_daily!G, month range)``
  N      Pre-Fee PnL                                    ``=SUM(H:I)-SUM(K:M)``
  P      Fund Fees / Mgmt                               ``SUMPRODUCT(text-month match × DAILY_CALC!C)``
  Q      Fund Fees / Incentive                          ``SUMPRODUCT(text-month match × DAILY_CALC!D)``
  R      Total                                          ``=P+Q``
  T      Net PnL                                        ``=N-R``
  V      EOM trading date (hidden helper)               **Python value** (last ``daily`` date in month)
  =====  =============================================  ============================================

  Columns C, E, J, O, S, U are intentionally left blank (template has no
  values/formulas there). The hidden helper column **V** is written as a
  precomputed Python date value so EOM ``SUMIFS`` matches exactly without
  depending on ``MAXIFS`` availability or implicit-intersection quirks
  (``@MAXIFS``) which can silently zero the column.
* **Borrow in Excel (``ALL_PAIRS``) vs Python rollups:** with ``inject_leg_pnl_borrow``,
  daily borrow in the workbook is the **per-row** short-leg accrual in
  ``_inject_all_pairs_leg_pnl_borrow_excel``. **Before** that, Python runs
  :func:`reallocate_net_underlying_borrow_by_under` so **aggregated** borrow in
  ``_daily_pair_rollups`` and ``pairwise_daily_net_pnl`` can *differ* from
  the sum of Excel row values when multiple pair rows share an underlying. See
  the docstring of that function; there is one economics path for the engine, one
  for transparent Excel per-leg formulas.

**Template:** ``resolve_template_xlsx()`` picks the first existing path among
``DC_ETF_PAIRWISE_TEMPLATE_XLSX``, ``notebooks/data/backtest/templates/DC_ETF_Arb_Pairwise_Template.xlsx``,
``~/Downloads/DC ETF Arb Pairwise Backtest Attribution.xlsx``, and ``~/Downloads/DC ETF Arb Attribution.xlsx``.
**Layout:** optional ``layout_golden_xlsx`` (or ``DC_ETF_PAIRWISE_LAYOUT_GOLDEN``) copies column widths / freeze panes
from a golden Pairwise workbook onto matching sheet names after data is written.
**Perf fee:** ``crystallize_trailing_partial_year=True`` (default) posts incentive on the last quarter-end
even when the series ends before year-end (e.g. asof 2026-03-31).

**Non-goals:** Do not rescale returns to match the fund; do not remove ``Monthly`` /
``Daily`` sheets.
================================================================================
"""

from __future__ import annotations

import argparse
import os
import shutil
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from scripts.daily_pair_workbook_formulas import (
    apply_portfolio_level_cost_model,
    build_book_daily_dataframe,
    inject_daily_pair_workbook_formulas,
    normalize_series_for_book_merge,
    reallocate_net_underlying_borrow_by_under,
)
from scripts.export_diamond_creek_v15_pair_ledger import _all_pairs_columns, _prep_pair_daily
from scripts.lp_fees_v15 import build_lp_fee_daily_cashflow_usd


def _repo_root() -> Path:
    return Path(__file__).resolve().parent.parent


def resolve_template_xlsx(explicit: Path | str | None) -> Path:
    """
    Resolve the workbook to copy before adding data sheets.

    Search order when ``explicit`` is ``None``:

    1. ``DC_ETF_PAIRWISE_TEMPLATE_XLSX`` environment variable
    2. ``notebooks/data/backtest/templates/DC_ETF_Arb_Pairwise_Template.xlsx``
    3. ``~/Downloads/DC ETF Arb Pairwise Backtest Attribution.xlsx`` (full layout)
    4. ``~/Downloads/DC ETF Arb Attribution.xlsx`` (2-sheet fund template)
    """
    if explicit is not None:
        p = Path(explicit)
        if p.is_file():
            return p
        raise FileNotFoundError(f"Template workbook not found: {p}")
    ev = os.environ.get("DC_ETF_PAIRWISE_TEMPLATE_XLSX")
    if ev:
        p = Path(ev)
        if p.is_file():
            return p
    home = Path.home()
    candidates = [
        _repo_root()
        / "notebooks"
        / "data"
        / "backtest"
        / "templates"
        / "DC_ETF_Arb_Pairwise_Template.xlsx",
        home / "Downloads" / "DC ETF Arb Pairwise Backtest Attribution.xlsx",
        home / "Downloads" / "DC ETF Arb Attribution.xlsx",
    ]
    for p in candidates:
        if p.is_file():
            return p
    raise FileNotFoundError(
        "No template workbook found. Place one of:\n"
        "  - notebooks/data/backtest/templates/DC_ETF_Arb_Pairwise_Template.xlsx\n"
        "  - ~/Downloads/DC ETF Arb Pairwise Backtest Attribution.xlsx\n"
        "  - ~/Downloads/DC ETF Arb Attribution.xlsx\n"
        "or set DC_ETF_PAIRWISE_TEMPLATE_XLSX to an .xlsx path."
    )


def resolve_layout_golden_xlsx(explicit: Path | str | None) -> Path | None:
    """
    Optional workbook whose column widths / freeze panes are copied onto the
    output for matching a golden Pairwise layout.

    When ``explicit`` is ``None``: use ``DC_ETF_PAIRWISE_LAYOUT_GOLDEN`` if set,
    else ``~/Downloads/DC ETF Arb Pairwise Backtest Attribution.xlsx`` if it exists.
    """
    if explicit is not None:
        p = Path(explicit)
        return p if p.is_file() else None
    ev = os.environ.get("DC_ETF_PAIRWISE_LAYOUT_GOLDEN")
    if ev:
        p = Path(ev)
        if p.is_file():
            return p
    p = Path.home() / "Downloads" / "DC ETF Arb Pairwise Backtest Attribution.xlsx"
    return p if p.is_file() else None


def _apply_column_layout_from_golden(target_wb, golden_path: Path) -> None:
    """Copy column widths, hidden flags, and freeze_panes from *golden* where sheet names match."""
    gwb = load_workbook(golden_path, data_only=False)
    try:
        for name in target_wb.sheetnames:
            if name not in gwb.sheetnames:
                continue
            src = gwb[name]
            dst = target_wb[name]
            for col_letter, dim in src.column_dimensions.items():
                if dim is None:
                    continue
                ddst = dst.column_dimensions[col_letter]
                if dim.width is not None:
                    ddst.width = dim.width
                if getattr(dim, "hidden", False):
                    ddst.hidden = True
            if src.freeze_panes is not None:
                dst.freeze_panes = src.freeze_panes
    finally:
        gwb.close()


def _coerce_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def _daily_pair_rollups(pair_daily: pd.DataFrame) -> pd.DataFrame:
    """One row per date: components for reconstructed pair net and active count."""
    p = pair_daily.copy()
    p["date"] = pd.to_datetime(p["date"], errors="coerce").dt.normalize()
    for c in (
        "daily_long_pnl_usd",
        "daily_short_pnl_usd",
        "daily_borrow_cost_usd",
        "daily_underlying_borrow_cost_usd",
        "long_notional_usd",
        "short_notional_usd",
        "long_sh",
        "short_sh",
        "underlying_price",
        "etf_price",
        "fed_funds_rate",
    ):
        if c not in p.columns:
            p[c] = np.nan
    lsh = _coerce_num(p["long_sh"]).fillna(0.0)
    ssh = _coerce_num(p["short_sh"]).fillna(0.0)
    pu = _coerce_num(p["underlying_price"]).fillna(0.0)
    pe = _coerce_num(p["etf_price"]).fillna(0.0)
    ln_file = _coerce_num(p["long_notional_usd"])
    sn_file = _coerce_num(p["short_notional_usd"])
    ln = ln_file.where(ln_file.notna(), lsh * pu)
    sn = sn_file.where(sn_file.notna(), ssh * pe)
    p["short_notional_abs_usd"] = sn.abs()
    p = p.assign(
        daily_long_pnl_usd=_coerce_num(p["daily_long_pnl_usd"]),
        daily_short_pnl_usd=_coerce_num(p["daily_short_pnl_usd"]),
        daily_borrow_cost_usd=_coerce_num(p["daily_borrow_cost_usd"]),
        daily_underlying_borrow_cost_usd=_coerce_num(p["daily_underlying_borrow_cost_usd"]),
        long_notional_usd=ln,
        short_notional_usd=sn,
        fed_funds_rate=_coerce_num(p["fed_funds_rate"]),
    )
    g = (
        p.groupby("date", sort=True)
        .agg(
            n_pairs=("etf", "count"),
            sum_long_pnl_usd=("daily_long_pnl_usd", "sum"),
            sum_short_pnl_usd=("daily_short_pnl_usd", "sum"),
            sum_borrow_usd=("daily_borrow_cost_usd", "sum"),
            sum_underlying_borrow_usd=("daily_underlying_borrow_cost_usd", "sum"),
            sum_long_notional=("long_notional_usd", "sum"),
            sum_short_notional_abs=("short_notional_abs_usd", "sum"),
            mean_fed_funds=("fed_funds_rate", "mean"),
        )
        .reset_index()
    )
    g["sum_gross_trading_usd"] = g["sum_long_pnl_usd"] + g["sum_short_pnl_usd"]
    g["pairwise_daily_net_pnl_usd"] = g["sum_gross_trading_usd"] - g["sum_borrow_usd"] - g["sum_underlying_borrow_usd"]
    return g


def _eom_notional_sums(pdw: pd.DataFrame, d0: pd.Timestamp) -> tuple[float, float, float]:
    """
    EOM snapshot: sums of signed / gross pair notionals (all rows for ``d0``).
    """
    t = pd.to_datetime(pdw["date"], errors="coerce").dt.normalize()
    s = pdw[t == pd.Timestamp(d0).normalize()]
    if s.empty:
        return 0.0, 0.0, 0.0
    ln = float(pd.to_numeric(s["long_notional_usd"], errors="coerce").fillna(0.0).sum())
    sn = float(pd.to_numeric(s["short_notional_usd"], errors="coerce").fillna(0.0).sum())
    if "gross_notional_usd" in s.columns:
        g = pd.to_numeric(s["gross_notional_usd"], errors="coerce")
        if g.notna().any():
            return ln, sn, float(g.fillna(0.0).sum())
    lsh = pd.to_numeric(s["long_sh"], errors="coerce").fillna(0.0)
    ssh = pd.to_numeric(s["short_sh"], errors="coerce").fillna(0.0)
    pu = pd.to_numeric(s["underlying_price"], errors="coerce").fillna(0.0)
    pe = pd.to_numeric(s["etf_price"], errors="coerce").fillna(0.0)
    gr = (np.abs(lsh) * pu + np.abs(ssh) * pe).sum()
    return ln, sn, float(gr)


def _eom_v15_monthly_books(pdw: pd.DataFrame, d0: pd.Timestamp) -> tuple[float, float]:
    """
    EOM snapshot for **Monthly E & G** (long book & short financing book), not
    signed ``long_notional_usd``/``short_notional_usd``. Sums ``long_margin_basis_usd``
    and ``short_financing_basis_usd`` (see module docstring and
    ``_inject_all_pairs_formulas``).
    """
    t = pd.to_datetime(pdw["date"], errors="coerce").dt.normalize()
    s = pdw[t == pd.Timestamp(d0).normalize()]
    if s.empty:
        return 0.0, 0.0
    if "long_margin_basis_usd" not in s.columns or "short_financing_basis_usd" not in s.columns:
        return 0.0, 0.0
    a = _coerce_num(s["long_margin_basis_usd"]).fillna(0.0)
    b = _coerce_num(s["short_financing_basis_usd"]).fillna(0.0)
    return float(a.sum()), float(b.sum())


def _xlsx_header_map(ws, row: int = 1) -> dict[str, int]:
    out: dict[str, int] = {}
    for c in ws[row]:
        v = c.value
        if v is None or str(v).strip() == "":
            continue
        out[str(v).strip()] = c.column
    return out


def _col_let(h: dict[str, int], name: str) -> str:
    c = h.get(name)
    if not c:
        raise KeyError(f"ALL_PAIRS missing {name!r} (export_dc_etf_arb_pairwise_workbook)")
    return get_column_letter(c)


def _book_raw_from_bt(
    bt: pd.DataFrame, dates: np.ndarray, *, attribution_base_capital: float
) -> pd.DataFrame:
    """
    One row per date: raw backtest columns for ``book_raw`` (values only) plus
    same-day *previous* cums and ``prev_nav`` (matches ``build_book_daily_dataframe``).
    """
    b = bt.sort_index().copy()
    b.index = pd.to_datetime(b.index, errors="coerce")
    if b.index.tz is not None:
        b.index = b.index.tz_convert("UTC").tz_localize(None)
    b.index = b.index.normalize()
    base = float(attribution_base_capital)
    rows: list[dict] = []
    for d in dates:
        d = pd.Timestamp(d).normalize()
        if d not in b.index:
            continue
        loc = b.index.get_loc(d)
        if isinstance(loc, slice):
            loc = int(loc.start) if loc.start is not None else 0
        elif isinstance(loc, np.ndarray):
            loc = int(loc.flat[0])
        else:
            loc = int(loc)
        r0 = b.iloc[loc]
        pre = b.iloc[loc - 1] if loc > 0 else None
        prev_cum = float(pd.to_numeric(pre.get("cum_costs"), errors="coerce") or 0.0) if pre is not None else 0.0
        prev_mdeb = float(
            (pd.to_numeric(pre.get("cum_margin_debit"), errors="coerce") or 0.0) if pre is not None else 0.0
        )
        prev_mcre = float(
            (pd.to_numeric(pre.get("cum_margin_credit"), errors="coerce") or 0.0) if pre is not None else 0.0
        )
        pnav = float(pd.to_numeric(pre.get("nav"), errors="coerce") or 0.0) if pre is not None else base
        rows.append(
            {
                "date": d,
                "nav": float(pd.to_numeric(r0.get("nav"), errors="coerce") or 0.0),
                "cum_costs": float(pd.to_numeric(r0.get("cum_costs"), errors="coerce") or 0.0),
                "prev_cum_costs": float(prev_cum),
                "daily_borrow": float(pd.to_numeric(r0.get("daily_borrow"), errors="coerce") or 0.0),
                "daily_long_pnl": float(pd.to_numeric(r0.get("daily_long_pnl"), errors="coerce") or 0.0),
                "daily_short_pnl": float(pd.to_numeric(r0.get("daily_short_pnl"), errors="coerce") or 0.0),
                "cum_margin_debit": float(pd.to_numeric(r0.get("cum_margin_debit"), errors="coerce") or 0.0),
                "cum_margin_credit": float(pd.to_numeric(r0.get("cum_margin_credit"), errors="coerce") or 0.0),
                "prev_cum_margin_debit": float(prev_mdeb),
                "prev_cum_margin_credit": float(prev_mcre),
                "prev_nav": float(pnav),
            }
        )
    return pd.DataFrame(rows)


def _lp_fee_row_formulas(row: int, *, daily_pnl_b_row: int) -> dict[str, str]:
    """
    Per-row ``LP_FEE`` formula bodies for **A** (date link to ``Daily PnL``) and **B**
    (quarterly management). **D / E / F** are filled with **Python values**, not
    formulas, by :func:`_lp_fee_row_values` so that the management fee fires
    reliably across Excel builds.

    Mgmt B = ``IF(include_fees_off, 0, IF(A=F, (mgmt_q)*E, 0))``.
    """
    a = f"A{row}"
    bdp = f"B{daily_pnl_b_row}"
    b = (
        f"=IF(OR(dc_pairwise_params!$B$2=0,dc_pairwise_params!$B$3=0),0,"
        f"IF({a}=F{row},(dc_pairwise_params!$B$3/4)*E{row},0))"
    )
    return {
        "A": f"='Daily PnL'!{bdp}",
        "B": b,
    }


def _lp_fee_helper_values(
    daily_dates: pd.DatetimeIndex,
    book_raw: pd.DataFrame,
    *,
    attribution_base_capital: float,
) -> pd.DataFrame:
    """
    Per-date precomputed LP_FEE helpers: q_first_trd, prev-quarter close NAV
    (mgmt base), q_last_trd. Index aligns 1:1 with ``daily_dates``.
    """
    dx = pd.DatetimeIndex(pd.to_datetime(daily_dates, errors="coerce")).normalize()
    qper = dx.to_period("Q")
    s = pd.Series(dx, index=qper)
    q_first = s.groupby(level=0).min()
    q_last = s.groupby(level=0).max()
    nav = (
        book_raw.set_index(pd.DatetimeIndex(pd.to_datetime(book_raw["date"], errors="coerce")).normalize())[
            "nav"
        ]
        .sort_index()
    )

    def _prev_close(d: pd.Timestamp) -> float:
        prev = nav.loc[nav.index < pd.Timestamp(d).normalize()]
        if prev.empty:
            return float(attribution_base_capital)
        return float(prev.iloc[-1])

    nav_base = {p: _prev_close(q_first.loc[p]) for p in q_first.index}
    rows = []
    for i, d in enumerate(dx):
        p = qper[i]
        rows.append(
            {
                "q_first_trd": q_first.loc[p].to_pydatetime(),
                "nav_base_mgmt": nav_base[p],
                "q_last_trd": q_last.loc[p].to_pydatetime(),
            }
        )
    return pd.DataFrame(rows)


def _eomax_trading_date_all_pairs(r: int) -> str:
    """
    Last ``ALL_PAIRS`` **trading** date in the month of ``B{r}`` (``yyyy-mm`` in B).

    Uses ``MAXIFS`` (Excel 2016+ / Microsoft 365) for reliable EOM = last row date
    in the month, so **SUMIFS(..., A, C)** in ``Monthly Attribution`` matches
    ``ALL_PAIRS!A`` when column **C** references this EOM. ``MAX( IF( … ) )`` can
    evaluate to 0 in some client builds when the month string or date serials
    do not line up, which leaves notionals and fee rollups at zero.
    """
    be = f"B{r}"
    a = "'ALL_PAIRS'!$A$2:$A$200000"
    return (
        f"IFERROR("
        f"MAXIFS({a},"
        f"{a},\">=\"&DATE(LEFT({be},4),MID({be},6,2),1),"
        f"{a},\"<=\"&EOMONTH(DATE(LEFT({be},4),MID({be},6,2),1),0)),0)"
    )


def _export_dc_etf_excel_bodies(
    out_path: Path,
    pd_work: pd.DataFrame,
    book: pd.DataFrame,
    book_raw: pd.DataFrame,
    fee_daily: pd.DataFrame,
    daily: pd.DataFrame,
    msum: pd.DataFrame,
    *,
    daily_pnl_source: str,
    attribution_base_capital: float,
    management_fee_rate_annual: float,
    incentive_fee_rate: float,
    fee_daycount: float,
    include_fees_in_daily_pnl: bool,
    trading_days: float,
    layout_golden_xlsx: Path | None = None,
) -> None:
    """Add ``ALL_PAIRS``, ``book_raw`` (values), ``book_daily`` (formulas), ``LP_FEE``, inject, ``DAILY_CALC``, monthly."""
    ap_cols = _all_pairs_columns()
    for c in ap_cols:
        if c not in pd_work.columns:
            pd_work[c] = np.nan
    for c in ("excel_long_trades_usd", "excel_short_trades_usd"):
        if c not in pd_work.columns:
            pd_work[c] = 0.0
        if c not in ap_cols:
            ap_cols = ap_cols + [c]
    all_pairs = pd_work[ap_cols].copy()
    if all_pairs.empty:
        raise ValueError("ALL_PAIRS export: empty after column prep")
    all_pairs = all_pairs.sort_values(["under", "etf", "date"], kind="mergesort").reset_index(drop=True)
    n_ap = len(all_pairs)
    n_d = len(daily)

    wb = load_workbook(out_path)
    for s in (
        "ALL_PAIRS",
        "book_raw",
        "book_daily",
        "LP_FEE",
        "DAILY_CALC",
        "dc_workbook_settings",
        "dc_pairwise_params",
    ):
        if s in wb.sheetnames:
            wb.remove(wb[s])

    ws = wb.create_sheet("ALL_PAIRS", 0)
    for r in dataframe_to_rows(all_pairs, index=False, header=True):
        ws.append(r)
    wbr = wb.create_sheet("book_raw", 1)
    for r in dataframe_to_rows(book_raw, index=False, header=True):
        wbr.append(r)
    wsb = wb.create_sheet("book_daily", 2)
    wsb.append(list(book.columns))
    n_book = len(book)
    for i in range(n_book):
        er = 2 + i
        wsb.cell(er, 1, f"=book_raw!A{er}")
        wsb.cell(er, 2, f"=book_raw!B{er}")
        wsb.cell(er, 3, f"=book_raw!C{er}-book_raw!D{er}")
        wsb.cell(er, 4, f"=B{er}-book_raw!L{er}")
        wsb.cell(er, 5, f"=book_raw!E{er}")
        wsb.cell(er, 6, f"=book_raw!F{er}+book_raw!G{er}")
        wsb.cell(
            er,
            7,
            f"=(book_raw!H{er}-book_raw!J{er})-(book_raw!I{er}-book_raw!K{er})",
        )
    wlp = wb.create_sheet("LP_FEE", 3)
    wlp.append(["date", "mgmt_usd", "perf_usd", "q_first_trd", "nav_base_mgmt", "q_last_trd"])
    dix2 = pd.DatetimeIndex(pd.to_datetime(daily["date"], errors="coerce")).normalize()
    fee_l = fee_daily.reindex(dix2, fill_value=0.0)
    for col in (4, 5, 6):
        wlp.column_dimensions[get_column_letter(col)].hidden = True
    wst = wb.create_sheet("dc_workbook_settings", 4)
    wst["A1"] = "trading_days"
    wst["B1"] = float(trading_days)
    wsp = wb.create_sheet("dc_pairwise_params", 5)
    wsp["A1"] = "daily_pnl_source"
    wsp["B1"] = str(daily_pnl_source)
    wsp["A2"] = "include_fees"
    wsp["B2"] = 0.0
    if include_fees_in_daily_pnl and (float(management_fee_rate_annual) > 0.0 or float(incentive_fee_rate) > 0.0):
        wsp["B2"] = 1.0
    wsp["A3"] = "management_fee_annual"
    wsp["B3"] = float(management_fee_rate_annual)
    wsp["A4"] = "attribution_base_capital"
    wsp["B4"] = float(attribution_base_capital)
    wsp["A5"] = "incentive_fee_rate"
    wsp["B5"] = float(incentive_fee_rate)
    wsp["A6"] = "fee_daycount"
    wsp["B6"] = float(fee_daycount)
    wb.save(out_path)
    wb.close()

    inject_daily_pair_workbook_formulas(
        out_path,
        all_pairs_data_rows=n_ap,
        portfolio_data_rows=0,
        checks_data_rows=0,
        inject_per_pair_sumifs=False,
        inject_leg_pnl_borrow=True,
        zero_pair_cost_columns=False,
        verbose=False,
        financing_daycount=360.0,
        include_portfolio_ledger_dc=False,
    )

    wb = load_workbook(out_path, data_only=False)
    h_ap = _xlsx_header_map(wb["ALL_PAIRS"], 1)
    c_net = _col_let(h_ap, "daily_pair_net_pnl_usd")
    c_dlp = _col_let(h_ap, "daily_long_pnl_usd")
    c_dsp = _col_let(h_ap, "daily_short_pnl_usd")
    c_dbor = _col_let(h_ap, "daily_borrow_cost_usd")
    c_fed = _col_let(h_ap, "fed_funds_rate")
    c_ln = _col_let(h_ap, "long_notional_usd")
    c_sn = _col_let(h_ap, "short_notional_usd")
    c_gr = _col_let(h_ap, "gross_notional_usd")

    a_last = 1 + n_d
    ws_d = wb["Daily PnL"]
    if ws_d.max_row >= 3:
        ws_d.delete_rows(3, ws_d.max_row - 2)
    for i in range(n_d):
        ws_d.cell(3 + i, 2, daily["date"].iloc[i])
    dix2 = pd.DatetimeIndex(pd.to_datetime(daily["date"], errors="coerce")).normalize()
    fee_l2 = fee_daily.reindex(dix2, fill_value=0.0)
    wlp = wb["LP_FEE"]
    helpers = _lp_fee_helper_values(
        dix2, book_raw, attribution_base_capital=float(attribution_base_capital)
    )
    for i in range(n_d):
        r = 2 + i
        bdp = 3 + i
        fm = _lp_fee_row_formulas(r, daily_pnl_b_row=bdp)
        wlp.cell(r, 1, fm["A"])
        wlp.cell(r, 2, fm["B"])
        wlp.cell(r, 3, float(fee_l2["perf_usd"].iloc[i]))
        wlp.cell(r, 4, helpers["q_first_trd"].iloc[i])
        wlp.cell(r, 5, float(helpers["nav_base_mgmt"].iloc[i]))
        wlp.cell(r, 6, helpers["q_last_trd"].iloc[i])

    wdc = wb.create_sheet("DAILY_CALC", len(wb.sheetnames))
    wdc["A1"] = "date_ref"
    wdc["B1"] = "pre_fee_pnl"
    wdc["C1"] = "management_fee"
    wdc["D1"] = "incentive_alloc"
    wdc["E1"] = "net_pnl"
    for i in range(n_d):
        rr = 2 + i
        wdc[f"A{rr}"] = f"='Daily PnL'!B{3 + i}"
        wdc[f"B{rr}"] = (
            f"=IF(dc_pairwise_params!$B$1=\"book_nav_change_usd\",IFERROR(VLOOKUP("
            f"A{rr},book_daily!$A:$D,4,FALSE),0),SUMIFS('ALL_PAIRS'!${c_net}$2:"
            f"${c_net}$600000,'ALL_PAIRS'!$A$2:$A$600000,A{rr}))"
        )
        wdc[f"C{rr}"] = (
            f"=IF(OR(dc_pairwise_params!$B$2=0,dc_pairwise_params!$B$3=0),0,"
            f"IFERROR(VLOOKUP(A{rr},LP_FEE!$A:$C,2,FALSE),0))"
        )
        wdc[f"D{rr}"] = (
            f"=IF(OR(dc_pairwise_params!$B$2=0,dc_pairwise_params!$B$5=0),0,"
            f"IFERROR(VLOOKUP(A{rr},LP_FEE!$A:$C,3,FALSE),0))"
        )
        wdc[f"E{rr}"] = f"=B{rr}-C{rr}-D{rr}"
    for i in range(n_d):
        r = 3 + i
        ws_d[f"C{r}"] = f"=DAILY_CALC!E{2 + i}"
        ws_d[f"D{r}"] = f"=COUNTIF('ALL_PAIRS'!$A:$A,B{r})"
    wsm = wb["Monthly Attribution"]
    if wsm.max_row >= 4:
        wsm.delete_rows(4, wsm.max_row - 3)
    while wsm.max_row > 3 and wsm.cell(wsm.max_row, 2).value in (None, ""):
        wsm.delete_rows(wsm.max_row)
    wsm.column_dimensions["V"].hidden = True

    eom_by_period = (
        daily.assign(_ym=daily["date"].dt.to_period("M"))
        .groupby("_ym", sort=True)["date"]
        .max()
    )

    def _month_range(r: int) -> tuple[str, str]:
        start = f"DATE(LEFT(B{r},4),MID(B{r},6,2),1)"
        end = f"EOMONTH(DATE(LEFT(B{r},4),MID(B{r},6,2),1),0)"
        return start, end

    def _sumifs_range(col_letter: str, sheet: str, r: int) -> str:
        s, e = _month_range(r)
        return (
            f"=SUMIFS('{sheet}'!${col_letter}$2:${col_letter}$200000,"
            f"'{sheet}'!$A$2:$A$200000,\">=\"&{s},"
            f"'{sheet}'!$A$2:$A$200000,\"<=\"&{e})"
        )

    for j, mrow in enumerate(msum.itertuples(index=False)):
        r = 4 + j
        s_dt, e_dt = _month_range(r)
        v_eom = f"$V{r}"
        wsm.cell(r, 2, mrow.month_str)
        eom_d = eom_by_period.loc[msum.index[j]]
        wsm.cell(r, 22, pd.Timestamp(eom_d).to_pydatetime())
        wsm[f"D{r}"] = (
            f"=IFERROR(AVERAGEIFS('ALL_PAIRS'!${c_fed}$2:${c_fed}$200000,"
            f"'ALL_PAIRS'!$A$2:$A$200000,\">=\"&{s_dt},"
            f"'ALL_PAIRS'!$A$2:$A$200000,\"<=\"&{e_dt}),0)"
        )
        wsm[f"F{r}"] = (
            f"=SUMIFS('ALL_PAIRS'!${c_ln}$2:${c_ln}$200000,"
            f"'ALL_PAIRS'!$A$2:$A$200000,{v_eom})"
        )
        wsm[f"G{r}"] = (
            f"=SUMIFS('ALL_PAIRS'!${c_sn}$2:${c_sn}$200000,"
            f"'ALL_PAIRS'!$A$2:$A$200000,{v_eom})"
        )
        wsm[f"H{r}"] = _sumifs_range(c_dlp, "ALL_PAIRS", r)
        wsm[f"I{r}"] = _sumifs_range(c_dsp, "ALL_PAIRS", r)
        wsm[f"K{r}"] = _sumifs_range("C", "book_daily", r)
        wsm[f"L{r}"] = _sumifs_range(c_dbor, "ALL_PAIRS", r)
        wsm[f"M{r}"] = _sumifs_range("G", "book_daily", r)
        wsm[f"N{r}"] = f"=SUM(H{r}:I{r})-SUM(K{r}:M{r})"
        wsm[f"P{r}"] = (
            f"=IF(OR(dc_pairwise_params!$B$2=0,dc_pairwise_params!$B$3=0),0,"
            f"SUMPRODUCT((TEXT(DAILY_CALC!$A$2:$A${a_last},\"yyyy-mm\")=B{r})*"
            f"DAILY_CALC!$C$2:$C${a_last}))"
        )
        wsm[f"Q{r}"] = (
            f"=IF(OR(dc_pairwise_params!$B$2=0,dc_pairwise_params!$B$5=0),0,"
            f"SUMPRODUCT((TEXT(DAILY_CALC!$A$2:$A${a_last},\"yyyy-mm\")=B{r})*"
            f"DAILY_CALC!$D$2:$D${a_last}))"
        )
        wsm[f"R{r}"] = f"=P{r}+Q{r}"
        wsm[f"T{r}"] = f"=N{r}-R{r}"
        for c_empty in (3, 5, 10, 15, 19, 21):
            wsm.cell(r, c_empty, None)
    if layout_golden_xlsx is not None:
        _apply_column_layout_from_golden(wb, layout_golden_xlsx)
    wb.save(out_path)
    wb.close()


def export_dc_etf_arb_pairwise_workbook(
    pair_daily: pd.DataFrame,
    bt: pd.DataFrame,
    out_xlsx: Path | str,
    *,
    template_xlsx: Path | str | None = None,
    attribution_base_capital: float,
    daily_pnl_source: str = "book_nav_change_usd",
    management_fee_rate_annual: float = 0.02,
    incentive_fee_rate: float = 0.20,
    fee_daycount: float = 365.0,
    include_fees_in_daily_pnl: bool = True,
    use_excel_formulas: bool = True,
    trading_days: float = 252.0,
    asof_end: str | pd.Timestamp | None = "2026-03-31",
    reallocate_underlying_borrow: bool = True,
    layout_golden_xlsx: Path | str | None = None,
    crystallize_trailing_partial_year: bool = True,
) -> Path:
    out_path = Path(out_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tpl = resolve_template_xlsx(template_xlsx)
    shutil.copy2(tpl, out_path)
    layout_golden = resolve_layout_golden_xlsx(layout_golden_xlsx)
    pd_in = pair_daily
    bt_in = bt
    if asof_end is not None:
        _end = pd.Timestamp(asof_end).normalize()
        dcol = pd.to_datetime(pd_in["date"], errors="coerce").dt.normalize()
        pd_in = pd_in.loc[dcol <= _end].copy()
        bidx = pd.DatetimeIndex(pd.to_datetime(bt_in.index, errors="coerce"))
        if bidx.tz is not None:
            bidx = bidx.tz_convert("UTC").tz_localize(None)
        bidx = bidx.normalize()
        bt_ = bt_in.copy()
        bt_.index = bidx
        bt_in = bt_[bt_.index <= _end]
    pd_work = _prep_pair_daily(pd_in)
    if pd_work.empty:
        raise ValueError("pair_daily is empty after prep (check asof_end filter)")
    if reallocate_underlying_borrow:
        pd_work = reallocate_net_underlying_borrow_by_under(pd_work, trading_days=trading_days)
    pd_work = apply_portfolio_level_cost_model(
        pd_work, verify_identities=True, zero_pair_level_books=False
    )
    pd_work["date"] = normalize_series_for_book_merge(pd_work["date"])
    daily = _daily_pair_rollups(pd_work)
    book = build_book_daily_dataframe(
        bt_in,
        daily["date"].unique(),
        attribution_base_capital=attribution_base_capital,
    )
    book["date"] = pd.to_datetime(book["date"]).dt.normalize()
    daily = daily.merge(
        book[
            [
                "date",
                "nav",
                "book_nav_change_usd",
                "book_daily_txn_usd",
                "book_daily_net_margin_usd",
            ]
        ],
        on="date",
        how="left",
    )
    daily["book_daily_txn_usd"] = _coerce_num(daily["book_daily_txn_usd"]).fillna(0.0)
    daily["book_daily_net_margin_usd"] = _coerce_num(daily["book_daily_net_margin_usd"]).fillna(0.0)
    daily["book_nav_change_usd"] = _coerce_num(daily["book_nav_change_usd"]).fillna(0.0)
    daily["nav"] = _coerce_num(daily["nav"]).fillna(0.0)
    if daily_pnl_source not in {"book_nav_change_usd", "pairwise_daily_net_pnl_usd"}:
        raise ValueError("daily_pnl_source must be 'book_nav_change_usd' or 'pairwise_daily_net_pnl_usd'")
    daily["pre_fee_daily_pnl_usd"] = _coerce_num(daily[daily_pnl_source]).fillna(0.0)
    dix = pd.DatetimeIndex(pd.to_datetime(daily["date"], errors="coerce")).normalize()
    fee_df = build_lp_fee_daily_cashflow_usd(
        bt_in["nav"],
        dix,
        attribution_base_capital=float(attribution_base_capital),
        management_fee_annual=float(management_fee_rate_annual),
        incentive_fee=float(incentive_fee_rate),
        crystallize_trailing_partial_year=bool(crystallize_trailing_partial_year),
    )
    fee_a = fee_df.reindex(dix, fill_value=0.0)
    daily["management_fee_daily_usd"] = _coerce_num(fee_a["mgmt_usd"])
    daily["incentive_fee_daily_usd"] = _coerce_num(fee_a["perf_usd"])
    if not include_fees_in_daily_pnl:
        daily["management_fee_daily_usd"] = 0.0
        daily["incentive_fee_daily_usd"] = 0.0
    last_idx = daily.groupby(daily["date"].dt.to_period("M"), sort=True)["date"].idxmax()
    eom_rows = daily.loc[last_idx]
    if isinstance(eom_rows, pd.Series):
        eom_rows = eom_rows.to_frame().T
    eom = eom_rows.assign(ym=lambda x: x["date"].dt.to_period("M")).set_index("ym")
    msum = (
        daily.assign(ym=daily["date"].dt.to_period("M"))
        .groupby("ym", sort=True)
        .agg(
            sum_long_pnl_usd=("sum_long_pnl_usd", "sum"),
            sum_short_pnl_usd=("sum_short_pnl_usd", "sum"),
            sum_borrow_usd=("sum_borrow_usd", "sum"),
            sum_underlying_borrow_usd=("sum_underlying_borrow_usd", "sum"),
            mean_fed_funds=("mean_fed_funds", "mean"),
            sum_book_txn_usd=("book_daily_txn_usd", "sum"),
            sum_book_margin_usd=("book_daily_net_margin_usd", "sum"),
            sum_management_fee_usd=("management_fee_daily_usd", "sum"),
            incentive_fee_usd=("incentive_fee_daily_usd", "sum"),
        )
    )
    msum["month_str"] = msum.index.astype(str)
    msum["eom_long_book_usd"] = msum.index.map(
        lambda per: _eom_v15_monthly_books(pd_work, eom.loc[per, "date"])[0]
    )
    msum["eom_short_financing_book_usd"] = msum.index.map(
        lambda per: _eom_v15_monthly_books(pd_work, eom.loc[per, "date"])[1]
    )
    msum["eom_long_notional_usd"] = msum.index.map(
        lambda per: _eom_notional_sums(pd_work, eom.loc[per, "date"])[0]
    )
    msum["eom_short_notional_usd"] = msum.index.map(
        lambda per: _eom_notional_sums(pd_work, eom.loc[per, "date"])[1]
    )
    msum["eom_gross_notional_usd"] = msum.index.map(
        lambda per: _eom_notional_sums(pd_work, eom.loc[per, "date"])[2]
    )
    msum["pre_fee_monthly_pnl_usd"] = (
        msum["sum_long_pnl_usd"]
        + msum["sum_short_pnl_usd"]
        - msum["sum_book_txn_usd"]
        - msum["sum_borrow_usd"]
        - msum["sum_book_margin_usd"]
    )
    daily["net_daily_pnl_after_fees_usd"] = (
        daily["pre_fee_daily_pnl_usd"] - daily["management_fee_daily_usd"] - daily["incentive_fee_daily_usd"]
    )
    book_raw = _book_raw_from_bt(
        bt_in, book["date"].values, attribution_base_capital=float(attribution_base_capital)
    )

    if use_excel_formulas:
        _export_dc_etf_excel_bodies(
            out_path,
            pd_work,
            book,
            book_raw,
            fee_df,
            daily,
            msum,
            daily_pnl_source=daily_pnl_source,
            attribution_base_capital=attribution_base_capital,
            management_fee_rate_annual=management_fee_rate_annual,
            incentive_fee_rate=incentive_fee_rate,
            fee_daycount=fee_daycount,
            include_fees_in_daily_pnl=include_fees_in_daily_pnl,
            trading_days=trading_days,
            layout_golden_xlsx=layout_golden,
        )
        return out_path
    wb = load_workbook(out_path)
    if "Daily PnL" not in wb.sheetnames or "Monthly Attribution" not in wb.sheetnames:
        wb.close()
        raise ValueError("Template missing expected sheets")
    ws_d = wb["Daily PnL"]
    dr0 = ws_d.max_row
    if dr0 >= 3:
        ws_d.delete_rows(3, dr0 - 2)
    for i, row in enumerate(daily.itertuples(index=False), start=0):
        r = 3 + i
        ws_d.cell(r, 2, row.date)
        v = float(row.net_daily_pnl_after_fees_usd) if include_fees_in_daily_pnl else float(row.pre_fee_daily_pnl_usd)
        ws_d.cell(r, 3, v)
        ws_d.cell(r, 4, int(row.n_pairs))
    ws_m = wb["Monthly Attribution"]
    mr0 = ws_m.max_row
    if mr0 >= 4:
        ws_m.delete_rows(4, mr0 - 3)
    while ws_m.max_row > 3 and ws_m.cell(ws_m.max_row, 2).value in (None, ""):
        ws_m.delete_rows(ws_m.max_row)
    for i, row in enumerate(msum.itertuples(index=False), start=0):
        r = 4 + i
        _per = msum.index[i]
        ws_m.cell(r, 2, row.month_str)
        ws_m.cell(r, 4, float(row.mean_fed_funds) if np.isfinite(row.mean_fed_funds) else 0.0)
        ws_m.cell(r, 6, float(row.eom_long_notional_usd))
        ws_m.cell(r, 7, float(row.eom_short_notional_usd))
        ws_m.cell(r, 8, float(row.sum_long_pnl_usd))
        ws_m.cell(r, 9, float(row.sum_short_pnl_usd))
        ws_m.cell(r, 11, float(row.sum_book_txn_usd))
        ws_m.cell(r, 12, float(row.sum_borrow_usd))
        ws_m.cell(r, 13, float(row.sum_book_margin_usd))
        ws_m.cell(r, 14, float(row.pre_fee_monthly_pnl_usd))
        ws_m.cell(r, 16, float(row.sum_management_fee_usd))
        ws_m.cell(r, 17, float(row.incentive_fee_usd))
        ws_m[f"R{r}"] = f"=P{r}+Q{r}"
        ws_m[f"T{r}"] = f"=N{r}-R{r}"
    wb.save(out_path)
    wb.close()
    return out_path


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--pair-csv", type=Path, help="Optional CSV with pair_daily rows for quick CLI tests")
    p.add_argument("--bt-csv", type=Path, help="Optional CSV with bt index date + nav,cum_costs columns")
    p.add_argument(
        "--template-xlsx",
        type=Path,
        default=None,
        help="Template .xlsx to copy first; default: auto-resolve (see resolve_template_xlsx)",
    )
    p.add_argument(
        "--layout-golden",
        type=Path,
        default=None,
        help="Optional golden Pairwise workbook to copy column widths / freeze panes from",
    )
    p.add_argument(
        "--no-reallocate-underlying-borrow",
        action="store_true",
        help="Skip reallocate_net_underlying_borrow_by_under (closer Excel row-sum vs engine if multi-pair per under)",
    )
    p.add_argument(
        "--no-trailing-perf-crystallize",
        action="store_true",
        help="Legacy: suppress incentive fee on trailing partial calendar year",
    )
    p.add_argument(
        "--out-xlsx",
        type=Path,
        default=Path("notebooks/data/backtest/DC ETF Arb Pairwise Backtest Attribution.xlsx"),
    )
    p.add_argument("--attribution-base-capital", type=float, default=10_000_000.0)
    p.add_argument("--daily-pnl-source", type=str, default="book_nav_change_usd")
    p.add_argument("--management-fee-rate-annual", type=float, default=0.0)
    p.add_argument("--incentive-fee-rate", type=float, default=0.0)
    p.add_argument("--fee-daycount", type=float, default=365.0)
    p.add_argument("--no-fees-in-daily-pnl", action="store_true")
    p.add_argument(
        "--asof-end",
        type=str,
        default="2026-03-31",
        help="Last calendar date to include in pair data and backtest (default: 2026-03-31; use empty to disable)",
    )
    p.add_argument("--value-bodies", action="store_true", help="Write numeric Daily/Monthly bodies (not formulas)")
    args = p.parse_args()
    if args.pair_csv and args.bt_csv:
        pair_daily = pd.read_csv(args.pair_csv, parse_dates=["date"])
        bt = pd.read_csv(args.bt_csv, parse_dates=["date"]).set_index("date")
        export_dc_etf_arb_pairwise_workbook(
            pair_daily,
            bt,
            args.out_xlsx,
            template_xlsx=args.template_xlsx,
            attribution_base_capital=args.attribution_base_capital,
            daily_pnl_source=args.daily_pnl_source,
            management_fee_rate_annual=args.management_fee_rate_annual,
            incentive_fee_rate=args.incentive_fee_rate,
            fee_daycount=args.fee_daycount,
            include_fees_in_daily_pnl=(not args.no_fees_in_daily_pnl),
            use_excel_formulas=(not args.value_bodies),
            asof_end=(args.asof_end.strip() or None),
            reallocate_underlying_borrow=(not args.no_reallocate_underlying_borrow),
            layout_golden_xlsx=args.layout_golden,
            crystallize_trailing_partial_year=(not args.no_trailing_perf_crystallize),
        )
        print("Wrote", args.out_xlsx.resolve())
    else:
        p.print_help()
        print(
            "\nProvide --pair-csv and --bt-csv for a standalone run, "
            "or import export_dc_etf_arb_pairwise_workbook from a notebook."
        )


if __name__ == "__main__":
    main()
