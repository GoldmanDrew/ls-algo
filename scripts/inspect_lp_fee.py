"""Inspect LP_FEE sheet around 2026-03 and the engine fee timeline."""
from openpyxl import load_workbook
import datetime as dt

NEW_P = r"C:\Users\werdn\Downloads\DC ETF Arb Pairwise Backtest Attribution.xlsx"

wb = load_workbook(NEW_P, data_only=True)
ws = wb["LP_FEE"]
print("LP_FEE shape:", ws.max_row, "x", ws.max_column)
header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print("header:", header)

# Print last 90 rows
for r in range(max(2, ws.max_row - 90), ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    print(r, row)

print()
print("--- dc_pairwise_params ---")
ws2 = wb["dc_pairwise_params"]
for r in range(1, ws2.max_row + 1):
    print(r, [ws2.cell(row=r, column=c).value for c in range(1, ws2.max_column + 1)])

print()
print("--- dc_workbook_settings ---")
ws3 = wb["dc_workbook_settings"]
for r in range(1, min(ws3.max_row, 30) + 1):
    print(r, [ws3.cell(row=r, column=c).value for c in range(1, ws3.max_column + 1)])

# Quarter-end tally: count rows with non-zero perf
print()
print("--- non-zero LP_FEE rows ---")
for r in range(2, ws.max_row + 1):
    a = ws.cell(row=r, column=1).value  # date
    b = ws.cell(row=r, column=2).value  # mgmt
    c = ws.cell(row=r, column=3).value  # perf
    if (b and abs(b) > 0.01) or (c and abs(c) > 0.01):
        print(r, a, "mgmt", b, "perf", c)
