"""Confirm the new engine is operating inside the configured 4.5x-5.0x band."""
from openpyxl import load_workbook
import statistics as st
from collections import defaultdict

NEW_P = r"C:\Users\werdn\Downloads\DC ETF Arb Pairwise Backtest Attribution.xlsx"

wb = load_workbook(NEW_P, data_only=True)
ws = wb["ALL_PAIRS"]
print("ALL_PAIRS shape:", ws.max_row, "x", ws.max_column)

# header at row 1
header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
col = {h: i + 1 for i, h in enumerate(header)}
need = ["date", "long_notional_usd", "short_notional_usd", "is_rebal"]
print("required cols:", {k: col.get(k) for k in need})

# Roll up gross/long/short by date (sum across pairs)
agg = defaultdict(lambda: [0.0, 0.0, 0])  # long, |short|, is_rebal_any
for r in range(2, ws.max_row + 1):
    d = ws.cell(row=r, column=col["date"]).value
    if d is None:
        continue
    ln = ws.cell(row=r, column=col["long_notional_usd"]).value or 0
    sn = ws.cell(row=r, column=col["short_notional_usd"]).value or 0
    isr = ws.cell(row=r, column=col["is_rebal"]).value or 0
    agg[d][0] += ln
    agg[d][1] += abs(sn)
    agg[d][2] = max(agg[d][2], int(isr))

# NAV from book_daily
ws_b = wb["book_daily"]
nav_by_date = {}
for r in range(2, ws_b.max_row + 1):
    d = ws_b.cell(row=r, column=1).value
    n = ws_b.cell(row=r, column=2).value
    if d is not None and isinstance(n, (int, float)):
        nav_by_date[d] = float(n)

dates = sorted(agg.keys())
ratios = []
ratios_rebal = []
for d in dates:
    nav = nav_by_date.get(d, None)
    if not nav:
        continue
    long_, sh_abs, isr = agg[d]
    g = long_ + sh_abs
    ratios.append((d, long_/nav, sh_abs/nav, g/nav, isr))
    if isr:
        ratios_rebal.append((d, long_/nav, sh_abs/nav, g/nav))

print(f"Days: {len(ratios)}, rebal days: {len(ratios_rebal)}")

def stats(seq, label):
    if not seq: return
    print(f"  {label}: min={min(seq):.3f}  max={max(seq):.3f}  mean={sum(seq)/len(seq):.3f}  median={sorted(seq)[len(seq)//2]:.3f}")

print()
print("Daily gross/NAV (all days):")
stats([r[1] for r in ratios], "long/NAV")
stats([r[2] for r in ratios], "short/NAV")
stats([r[3] for r in ratios], "gross/NAV")

print()
print("Daily gross/NAV (rebal days only — should be tightly inside 4.5-5.0):")
stats([r[1] for r in ratios_rebal], "long/NAV")
stats([r[2] for r in ratios_rebal], "short/NAV")
stats([r[3] for r in ratios_rebal], "gross/NAV")

# Histogram bins for gross/NAV on all days
print()
print("All-day gross/NAV distribution:")
import math
bins = [3.0, 3.5, 4.0, 4.25, 4.5, 4.75, 5.0, 5.25, 5.5, 5.75, 6.0, 7.0]
counts = [0]*(len(bins)-1)
for r in ratios:
    g = r[3]
    for i in range(len(bins)-1):
        if bins[i] <= g < bins[i+1]:
            counts[i]+=1
            break
for i in range(len(bins)-1):
    bar = '#' * int(counts[i]/2)
    print(f"  [{bins[i]:.2f}, {bins[i+1]:.2f}): {counts[i]:>4} {bar}")

# By year, gross/NAV avg
print()
print("By year (rebal days only):")
import collections
year_groups = collections.defaultdict(list)
for d, l, s, g in ratios_rebal:
    year_groups[d.year].append(g)
for y, vs in sorted(year_groups.items()):
    print(f"  {y}: count={len(vs)}, mean={sum(vs)/len(vs):.3f}, max={max(vs):.3f}, min={min(vs):.3f}")
